import asyncio
import logging
import html as html_module
import traceback
import re
from io import BytesIO
from pathlib import Path
from datetime import datetime, timedelta
import pytz
from openpyxl import Workbook
from aiogram import Bot, Dispatcher, types
from aiogram.types import ErrorEvent
from aiogram.exceptions import TelegramBadRequest
from aiogram.client.default import DefaultBotProperties
from aiogram.types import Message, InlineKeyboardMarkup, InlineKeyboardButton, CallbackQuery, ReplyKeyboardRemove
from aiogram.types.input_file import BufferedInputFile
from aiogram.filters import Command

from utils import teacher_main_keyboard, create_language_selection_keyboard_for_user, create_subject_keyboard, cancel_keyboard, student_main_keyboard, create_language_selection_keyboard_for_self
from i18n import format_group_level_display, t, detect_lang_from_user, level_ui_label
from config import (
    TEACHER_BOT_TOKEN,
    ADMIN_BOT_TOKEN,
    STUDENT_BOT_TOKEN,
    ADMIN_CHAT_IDS,
    ALL_ADMIN_IDS,
    TEACHER_WEBHOOK_PORT,
)
from bot_runtime import create_resilient_bot, run_bot_dispatcher, spawn_guarded_task
from db import (
    get_groups_by_teacher, get_user_by_id, get_group_users, add_attendance, is_access_active, get_group,
    get_teacher_total_students, get_user_by_telegram, get_conn, add_feedback, get_dcoins,
    get_daily_test_upload_permission, get_daily_tests_stock_by_teacher, get_daily_test_history_for_teacher,
    get_staff_leaderboard_by_subject,
    get_staff_leaderboard_student_count,
    get_present_students_for_group_date,
    is_lesson_holiday,
    is_lesson_otmen_date_cancelled,
    get_all_teachers,
    get_groups_with_temporary_access_for_teacher,
    get_active_temporary_assignments_for_pair,
    create_temporary_group_assignment,
    get_active_temporary_assignments_by_owner,
    cancel_temporary_assignments_for_pair,
    teacher_has_temporary_group_access,
    create_arena_group_session, get_arena_group_session, set_arena_group_session_status,
    set_arena_group_session_expected_players,
    copy_daily_tests_bank_rows_to_arena_questions,
    ensure_subject_dcoin_schema,
    ensure_dcoin_schema_migrations,
    validate_dcoin_runtime_ready,
    get_arena_group_session_questions,
    set_arena_group_session_teacher_message,
    enqueue_arena_group_teacher_refresh,
    get_group_arena_teacher_snapshot,
    list_arena_group_session_answers_for_export,
    pop_arena_group_teacher_refresh_session,
)
from lesson_window import is_group_lesson_window_active
from auth import verify_login, activate_user, restore_sessions_on_startup, process_login_message, get_login_state, set_login_state, clear_login_state, AuthMiddleware
from attendance_manager import (
    attendance_scheduler,
    build_attendance_keyboard,
    get_panel_ui_lang,
    get_session,
    set_session_closed,
    set_session_opened,
    set_attendance_student_notify_bot,
)
from logging_config import get_logger
from bot_error_notify import notify_admins_unhandled_bot_error

# Setup logger
logger = get_logger(__name__)

bot: Bot | None = None
admin_bot: Bot | None = None
student_notify_bot: Bot | None = None
dp = Dispatcher()


@dp.errors()
async def teacher_global_error_handler(event: ErrorEvent):
    exc = event.exception
    if isinstance(exc, TelegramBadRequest):
        msg = str(exc).lower()
        if "query is too old" in msg or "query id is invalid" in msg:
            logger.warning("Suppressed stale callback error: %s", exc)
            return True
    try:
        await notify_admins_unhandled_bot_error(
            bot_label="Teacher bot",
            event=event,
            admin_bot_instance=admin_bot,
        )
    except Exception:
        logger.exception("Adminlarga teacher bot xato xabari yuborilmadi")
    return False


# Setup authentication middleware
auth_middleware = AuthMiddleware(bot_type='teacher', expected_login_type=3)
dp.message.middleware(auth_middleware)
dp.callback_query.middleware(auth_middleware)

# Teacher state for main menu
teacher_state = {}


def get_teacher_state(chat_id):
    return teacher_state.setdefault(chat_id, {'step': None, 'data': {}})


_LEVEL_STOCK_ORDER = ("A1", "A2", "B1", "B2", "C1", "MIXED")


def _teacher_format_daily_tests_stock_html(lang: str, teacher_id: int, subject: str) -> str:
    stock = get_daily_tests_stock_by_teacher(teacher_id, subject)

    def _lvl_sort_key(item):
        lvl = (item[0] or "").strip().upper()
        try:
            return (_LEVEL_STOCK_ORDER.index(lvl), lvl)
        except ValueError:
            return (99, lvl)

    lines = [
        t(lang, "teacher_daily_tests_stock_report_title", subject=html_module.escape(subject)),
        "",
    ]
    for lvl, cnt in sorted(stock["stock"].items(), key=_lvl_sort_key):
        lines.append(
            t(
                lang,
                "teacher_daily_tests_stock_level_line",
                level=html_module.escape(lvl),
                count=int(cnt),
            )
        )
    lines.append("")
    lines.append(t(lang, "daily_tests_stock_total", total=int(stock.get("total") or 0)))
    return "\n".join(lines)


def _teacher_format_daily_tests_history(lang: str, teacher_id: int) -> str:
    history = get_daily_test_history_for_teacher(teacher_id, days=14)
    lines = [t(lang, "teacher_daily_tests_history_title")]
    if not history:
        lines.append(t(lang, "teacher_daily_tests_history_empty"))
        return "\n".join(lines)
    avg_lbl = t(lang, "teacher_daily_tests_history_avg_label")
    for a in history:
        lines.append(
            t(
                lang,
                "teacher_daily_tests_history_row_line",
                td=a.get("test_date"),
                completed=a.get("completed_attempts") or 0,
                correct_total=a.get("correct_total") or 0,
                wrong_total=a.get("wrong_total") or 0,
                unanswered_total=a.get("unanswered_total") or 0,
                avg_label=avg_lbl,
                avg_net=a.get("avg_net_dcoins") or 0,
            )
        )
    return "\n".join(lines)


async def _teacher_send_dcoin_leaderboard(message: Message, lang: str, subject: str, page: int) -> None:
    per_page = 20
    total_users = get_staff_leaderboard_student_count(subject)
    total_pages = max(1, (total_users + per_page - 1) // per_page)
    page = max(0, min(int(page), total_pages - 1))
    offset = page * per_page
    rows = get_staff_leaderboard_by_subject(subject, limit=per_page, offset=offset)
    lines = [
        t(lang, "staff_dcoin_leaderboard_title", subject=html_module.escape(subject)),
        t(lang, "staff_dcoin_leaderboard_subtitle"),
        "",
    ]
    if not rows:
        lines.append(t(lang, "staff_dcoin_leaderboard_empty"))
    else:
        for i, r in enumerate(rows):
            raw_name = f"{r.get('first_name') or ''} {r.get('last_name') or ''}".strip() or "-"
            name = html_module.escape(raw_name)
            bal = float(r.get("dcoin_balance") or 0)
            rank = offset + i + 1
            lines.append(t(lang, "staff_dcoin_leaderboard_line", rank=rank, name=name, dcoin=bal))
    lines.append("")
    lines.append(
        t(
            lang,
            "staff_dcoin_leaderboard_footer",
            page=page + 1,
            total_pages=total_pages,
            total=total_users,
        )
    )
    nav = []
    if page > 0:
        nav.append(
            InlineKeyboardButton(
                text=t(lang, "btn_prev"),
                callback_data=f"tc_lb_p:{page - 1}",
            )
        )
    if page < total_pages - 1:
        nav.append(
            InlineKeyboardButton(
                text=t(lang, "btn_next"),
                callback_data=f"tc_lb_p:{page + 1}",
            )
        )
    kb = InlineKeyboardMarkup(inline_keyboard=[nav] if nav else [])
    await message.answer("\n".join(lines), parse_mode="HTML", reply_markup=kb)


def reset_teacher_state(chat_id):
    teacher_state.pop(chat_id, None)


LIST_PAGE_SIZE = 10
TASK_TZ = pytz.timezone("Asia/Tashkent")


def _teacher_can_manage_group(user: dict, group: dict | None) -> bool:
    if not user or not group:
        return False
    if group.get("teacher_id") == user.get("id"):
        return True
    return teacher_has_temporary_group_access(user.get("id"), group.get("id"))


def _has_lesson_today(g: dict) -> bool:
    now = datetime.now(TASK_TZ)
    today = now.strftime("%Y-%m-%d")
    weekday = now.weekday()
    ld = (g.get("lesson_date") or "").strip()
    if is_lesson_holiday(today):
        return False
    if not ld:
        return False
    if re.match(r"^\d{4}-\d{2}-\d{2}$", ld):
        return ld == today
    code = ld.upper()
    if code in ("MWF", "MON/WED/FRI", "MON,WED,FRI", "ODD"):
        return weekday in (0, 2, 4)
    if code in ("TTS", "TUE/THU/SAT", "TUE,THU,SAT", "EVEN"):
        return weekday in (1, 3, 5)
    return False


def _build_numbered_keyboard(
    items: list[dict],
    page: int,
    pick_cb_builder,
    page_cb_prefix: str,
    extra_rows: list[list[InlineKeyboardButton]] | None = None,
) -> InlineKeyboardMarkup:
    total = len(items)
    total_pages = max(1, (total + LIST_PAGE_SIZE - 1) // LIST_PAGE_SIZE)
    page = max(0, min(page, total_pages - 1))
    start = page * LIST_PAGE_SIZE
    sub = items[start:start + LIST_PAGE_SIZE]

    rows: list[list[InlineKeyboardButton]] = []
    nums = []
    for idx, item in enumerate(sub, start=1):
        nums.append(InlineKeyboardButton(text=str(idx), callback_data=pick_cb_builder(item)))
    if nums:
        rows.append(nums[:5])
        if len(nums) > 5:
            rows.append(nums[5:10])

    nav = []
    if page > 0:
        nav.append(InlineKeyboardButton(text="⬅️", callback_data=f"{page_cb_prefix}:{page-1}"))
    nav.append(InlineKeyboardButton(text=f"{page+1}/{total_pages}", callback_data="att_noop"))
    if page < total_pages - 1:
        nav.append(InlineKeyboardButton(text="➡️", callback_data=f"{page_cb_prefix}:{page+1}"))
    rows.append(nav)

    if extra_rows:
        rows.extend(extra_rows)
    return InlineKeyboardMarkup(inline_keyboard=rows)


def _teacher_name(u: dict) -> str:
    return f"{u.get('first_name','')} {u.get('last_name','')}".strip() or f"ID {u.get('id')}"


def _fmt_date(value: str | None) -> str:
    from i18n import format_date_display
    return format_date_display(value or "-")


def _manageable_groups_for_teacher(teacher_id: int) -> list[dict]:
    own = get_groups_by_teacher(teacher_id) or []
    temp = get_groups_with_temporary_access_for_teacher(teacher_id) or []
    merged: dict[int, dict] = {}
    for g in own + temp:
        merged[int(g["id"])] = g
    return list(merged.values())


def _format_group_items(lang: str, groups: list[dict], include_teacher: bool = True) -> str:
    lines = []
    for i, g in enumerate(groups, start=1):
        students_in_group = len(get_group_users(g["id"]))
        lines.append(f"{i}. Group name: {g.get('name','-')}")
        lines.append(f"   Group level: {format_group_level_display(lang, g.get('level'), subject=g.get('subject'))}")
        if include_teacher:
            teacher = get_user_by_id(g.get("teacher_id")) if g.get("teacher_id") else None
            lines.append(f"   Group teacher: {_teacher_name(teacher or {})}")
        lines.append(f"   Students in group: {students_in_group}")
    return "\n".join(lines) if lines else "-"


def _teacher_arena_groups_page(groups: list[dict], page: int, page_size: int = 10) -> list[dict]:
    start = max(0, page) * page_size
    return groups[start : start + page_size]


def _render_teacher_arena_group_list_text(lang: str, chunk: list[dict], page: int, total_pages: int) -> str:
    text = t(lang, "teacher_group_arena_list_title", page=page + 1, total=total_pages)
    for i, g in enumerate(chunk, start=1):
        students_count = len(get_group_users(g["id"]))
        level = format_group_level_display(lang, g.get("level"), subject=g.get("subject"))
        text += (
            f"\n<b>{i}. {html_module.escape(g.get('name') or '-')}</b>\n"
            f"   🎯 {html_module.escape(str(level))}\n"
            f"   👥 {students_count}\n"
        )
    return text


def _teacher_arena_group_pick_keyboard(
    page_groups: list[dict], page: int, total_pages: int, lang: str
) -> InlineKeyboardMarkup:
    rows = []
    nums = []
    for idx, g in enumerate(page_groups):
        nums.append(
            InlineKeyboardButton(
                text=str(idx + 1 + page * 10),
                callback_data=f"teacher_arena_grp:pick:{g['id']}",
            )
        )
    if nums:
        rows.append(nums[:5])
        if len(nums) > 5:
            rows.append(nums[5:10])
    nav = []
    if page > 0:
        nav.append(
            InlineKeyboardButton(text=t(lang, "btn_back"), callback_data=f"teacher_arena_grp:page:{page-1}")
        )
    if page < total_pages - 1:
        nav.append(
            InlineKeyboardButton(text=t(lang, "btn_next"), callback_data=f"teacher_arena_grp:page:{page+1}")
        )
    if nav:
        rows.append(nav)
    return InlineKeyboardMarkup(inline_keyboard=rows)


def _format_teacher_arena_live_text(snapshot: dict, lang: str) -> str:
    sn = snapshot.get("session") or {}
    present = snapshot.get("present") or []
    attempts = {int(a["user_id"]): a for a in snapshot.get("attempts") or []}
    lines = [t(lang, "teacher_group_arena_live_title", session_id=sn.get("id", "-"))]
    for u in present:
        uid = int(u["id"])
        name = f"{u.get('first_name','')} {u.get('last_name','')}".strip() or str(uid)
        a = attempts.get(uid)
        if not a:
            lines.append(f"⏳ {html_module.escape(name)} — {t(lang, 'teacher_group_arena_status_not_started')}")
        elif a.get("finished_at"):
            lines.append(
                f"✅ {html_module.escape(name)} — "
                f"{t(lang, 'teacher_group_arena_status_done', correct=a.get('correct'), wrong=a.get('wrong'), skipped=a.get('unanswered'))}"
            )
        else:
            lines.append(f"🕓 {html_module.escape(name)} — {t(lang, 'teacher_group_arena_status_in_progress')}")
    return "\n".join(lines)


def _build_group_arena_questions_xlsx(session_id: int) -> bytes:
    qs = get_arena_group_session_questions(session_id)
    wb = Workbook()
    ws = wb.active
    ws.title = "questions"
    ws.append(["#", "Type", "Question", "A", "B", "C", "D", "Correct_ix"])
    for i, q in enumerate(qs, start=1):
        ws.append(
            [
                i,
                q.get("question_type") or "",
                q.get("question") or "",
                q.get("option_a") or "",
                q.get("option_b") or "",
                q.get("option_c") or "",
                q.get("option_d") or "",
                q.get("correct_option_index"),
            ]
        )
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _build_group_arena_matrix_xlsx(session_id: int) -> bytes:
    qs = get_arena_group_session_questions(session_id)
    rows = list_arena_group_session_answers_for_export(session_id)
    wb = Workbook()
    ws = wb.active
    ws.title = "matrix"
    ws.append(["Student", "Q#", "Question", "Selected", "Correct", "Type"])
    q_by_order = {i: q for i, q in enumerate(qs, start=1)}
    for r in rows:
        ord_ = int(r["question_order"])
        q = q_by_order.get(ord_, {})
        sel = r.get("selected_option_index")
        if sel is None:
            sel_label = "—"
        else:
            try:
                si = int(sel)
                sel_label = ["A", "B", "C", "D"][si] if 0 <= si <= 3 else str(sel)
            except Exception:
                sel_label = str(sel)
        name = f"{r.get('first_name','')} {r.get('last_name','')}".strip()
        ws.append(
            [
                name,
                ord_,
                (q.get("question") or "")[:200],
                sel_label,
                "yes" if int(r.get("is_correct") or 0) else "no",
                q.get("question_type") or "",
            ]
        )
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


async def _refresh_teacher_arena_message(bot: Bot, session_id: int) -> None:
    sess = get_arena_group_session(session_id)
    if not sess:
        return
    chat_id = sess.get("teacher_chat_id")
    msg_id = sess.get("teacher_status_message_id")
    if not chat_id or not msg_id:
        return
    teacher = get_user_by_id(int(sess.get("created_by_teacher_id") or 0)) or {}
    lang = detect_lang_from_user(teacher)
    snap = get_group_arena_teacher_snapshot(session_id)
    if not snap:
        return
    text = _format_teacher_arena_live_text(snap, lang)
    completed = (sess.get("status") or "").strip() == "completed" or int(sess.get("rewards_distributed") or 0) == 1
    kb = None
    if completed:
        kb = InlineKeyboardMarkup(
            inline_keyboard=[
                [
                    InlineKeyboardButton(
                        text=t(lang, "teacher_group_arena_download_matrix_btn"),
                        callback_data=f"teacher_arena_matrix_dl:{session_id}",
                    )
                ]
            ]
        )
    try:
        await bot.edit_message_text(
            text,
            chat_id=int(chat_id),
            message_id=int(msg_id),
            parse_mode="HTML",
            reply_markup=kb,
        )
    except TelegramBadRequest:
        pass


async def teacher_arena_refresh_worker(bot: Bot) -> None:
    while True:
        try:
            await asyncio.sleep(3)
            sid = pop_arena_group_teacher_refresh_session()
            if not sid:
                continue
            await _refresh_teacher_arena_message(bot, sid)
        except Exception:
            logger.exception("teacher_arena_refresh_worker failed")
            await asyncio.sleep(2)


def _temp_assignment_meta(group: dict, lang: str, selected_lessons: list[dict]) -> str:
    group_name = group.get("name", "-")
    subject = group.get("subject") or "-"
    level = format_group_level_display(lang, group.get("level"), subject=group.get("subject"))
    days_raw = group.get("lesson_date") or "-"
    if str(days_raw).upper() in ("MWF", "MON/WED/FRI", "MON,WED,FRI"):
        days = t(lang, "admin_btn_lesson_days_mwf")
    elif str(days_raw).upper() in ("TTS", "TUE/THU/SAT", "TUE,THU,SAT"):
        days = t(lang, "admin_btn_lesson_days_tts")
    else:
        days = str(days_raw)
    start = (group.get("lesson_start") or "-")[:5]
    end = (group.get("lesson_end") or "-")[:5]
    schedule_lines = []
    for ls in selected_lessons:
        schedule_lines.append(f"- {_fmt_date(ls.get('lesson_date'))} ({ls.get('weekday')}) {ls.get('lesson_start')}-{ls.get('lesson_end')}")
    schedule_text = "\n".join(schedule_lines) if schedule_lines else "-"
    students_count = len(get_group_users(group.get("id")))
    return t(
        lang,
        "teacher_temp_assignment_meta",
        group_name=group_name,
        subject=subject,
        level=level,
        days=days,
        start=start,
        end=end,
        students_count=students_count,
        schedule_text=schedule_text,
    )


def _temp_assignment_admin_chat_targets(group: dict) -> set[int]:
    """
    Notify both:
    - group owner admin (groups.owner_admin_id -> users.telegram_id)
    - general admins (ADMIN_CHAT_IDS)
    """
    targets: set[int] = set()
    owner_admin_id = group.get("owner_admin_id")
    if owner_admin_id:
        owner_admin = get_user_by_id(owner_admin_id)
        if owner_admin and owner_admin.get("telegram_id"):
            try:
                targets.add(int(owner_admin["telegram_id"]))
            except Exception:
                logger.warning("Invalid owner admin telegram_id for owner_admin_id=%s", owner_admin_id)
    for aid in ADMIN_CHAT_IDS:
        try:
            targets.add(int(aid))
        except Exception:
            logger.warning("Invalid ADMIN_CHAT_IDS entry: %s", aid)
    return targets


def _compute_upcoming_lessons(group: dict, take: int = 30) -> list[dict]:
    raw = (group.get("lesson_date") or "").strip()
    start = (group.get("lesson_start") or "-")[:5]
    end = (group.get("lesson_end") or "-")[:5]
    now = datetime.now(TASK_TZ).date()
    out: list[dict] = []
    if re.match(r"^\d{4}-\d{2}-\d{2}$", raw):
        d = datetime.strptime(raw, "%Y-%m-%d").date()
        if d >= now:
            out.append({"lesson_date": raw, "lesson_start": start, "lesson_end": end, "weekday": d.strftime("%A")})
        return out

    code = raw.upper()
    week_days = []
    if code in ("MWF", "MON/WED/FRI", "MON,WED,FRI", "ODD"):
        week_days = [0, 2, 4]
    elif code in ("TTS", "TUE/THU/SAT", "TUE,THU,SAT", "EVEN"):
        week_days = [1, 3, 5]
    else:
        return out

    cursor = now
    while len(out) < take:
        if cursor.weekday() in week_days:
            out.append(
                {
                    "lesson_date": cursor.strftime("%Y-%m-%d"),
                    "lesson_start": start,
                    "lesson_end": end,
                    "weekday": cursor.strftime("%A"),
                }
            )
        cursor += timedelta(days=1)
    return out

@dp.message(Command('start'))
async def cmd_start(message: Message):
    """Handle /start command - check login status or show login instructions"""
    logger.debug(f"/start called by user_id={message.from_user.id}")
    telegram_id = str(message.from_user.id)
    user = get_user_by_telegram(telegram_id)
    
    # If user is already logged in, show main menu
    if user and user.get('logged_in'):
        lang = detect_lang_from_user(user)
        await message.answer(t(lang, 'welcome_back'), reply_markup=teacher_main_keyboard(lang))
        return
    
    # Clear any existing login state and start new login flow
    clear_login_state(telegram_id)
    reset_teacher_state(message.chat.id)
    
    # Start two-step login flow
    set_login_state(telegram_id, {'step': 'ask_login', 'data': {}})
    
    lang = detect_lang_from_user(message.from_user)
    
    await message.answer(t(lang, 'teacher_login_title'))
    await message.answer(t(lang, 'ask_login_id'))


@dp.message(lambda m: not (m.text or '').startswith('/'))
async def handle_login_and_messages(message: Message):
    """Ikki bosqichli login + eski formatni qo‘llab-quvvatlaydi"""
    telegram_id = str(message.from_user.id)
    login_state = get_login_state(telegram_id)

    # Login jarayonida bo‘lsa yoki eski formatda (:) bo‘lsa — process qil
    if login_state.get('step') in ('ask_login', 'ask_password') or ':' in (message.text or ''):
        success = await process_login_message(message, expected_login_type=3)
        if success:
            user = get_user_by_telegram(telegram_id)
            lang = detect_lang_from_user(user or message.from_user)
            await message.answer(t(lang, 'login_success'))
            await message.answer(t(lang, 'welcome_back'), reply_markup=teacher_main_keyboard(lang))
        return  # Jarayon tugamaguncha boshqa hech narsa qilma

    # Agar login tugagan bo‘lsa — oddiy authenticated handler
    await handle_authenticated_message(message)


async def handle_authenticated_message(message: Message):
    """Handle messages from authenticated teachers"""
    logger.info(f"💬 TEACHER MESSAGE: {message.text} | User: {message.from_user.id}")
    user = get_user_by_telegram(str(message.from_user.id))
    lang = detect_lang_from_user(user or message.from_user)
    
    logger.info(f"🔍 Teacher user data: {user}")
    
    if not user or user.get('login_type') != 3:
        await message.answer(t(lang, 'please_send_start'))
        return

    if user['blocked']:
        await message.answer(t(lang, 'blocked_contact_admin'))
        return

    ts = get_teacher_state(message.chat.id)
    if ts.get('step') == 'feedback_confirm':
        await message.answer(t(lang, 'feedback_confirm_use_buttons'))
        return
    if ts.get('step') == 'feedback_wait':
        await handle_teacher_feedback_text(message, user, lang)
        return
    if ts.get('step') == 'teacher_arena_ai_count':
        txt = (message.text or '').strip()
        if txt == t(lang, 'btn_cancel'):
            ts['step'] = None
            ts['data'] = {}
            await message.answer(t(lang, 'select_from_menu'), reply_markup=teacher_main_keyboard(lang))
            return

        if not txt.isdigit():
            await message.answer(t(lang, "validation_only_number"))
            return

        cnt = max(1, min(100, int(txt)))
        gid = int(ts['data'].get('arena_group_id') or 0)
        topic_mode = (ts['data'].get('topic_mode') or 'any').strip().lower()

        group = get_group(gid)
        if not group:
            await message.answer(t(lang, 'group_not_found'))
            ts['step'] = None
            return

        from auth import normalize_level_to_cefr
        from ai_generator import allowed_levels_for_ai_pipeline, generate_daily_tests_and_insert

        arena_subject = (group.get("subject") or user.get("subject") or "English").strip().title()
        arena_level = "MIXED" if topic_mode == "any" else normalize_level_to_cefr(group.get("level") or "A1")

        allowed_levels = allowed_levels_for_ai_pipeline(arena_subject)
        if arena_level not in allowed_levels:
            arena_level = "MIXED" if topic_mode == "any" else ("B1" if arena_subject == "Russian" else "A1")

        start_ts = datetime.utcnow()

        logger.info(
            "teacher_arena_ai_count start user_id=%s gid=%s cnt=%s subject=%s level=%s topic_mode=%s",
            user.get("id"),
            gid,
            cnt,
            arena_subject,
            arena_level,
            topic_mode,
        )

        progress_msg = await message.answer(t(lang, "teacher_group_arena_ai_started"))

        try:
            await generate_daily_tests_and_insert(
                subject=arena_subject,
                level=arena_level,
                count=cnt,
                created_by=user['id'],
            )
        except Exception as e:
            await message.answer(t(lang, "teacher_group_arena_generation_error", error=str(e)))
            ts['step'] = None
            return

        # === ENG ISHONCHLI TANLASH ===
        conn = get_conn()
        cur = conn.cursor()
        try:
            if topic_mode == "any":
                # In MIXED/any mode, generator may store Russian rows as B1/B2 (not literal MIXED).
                # So we MUST NOT filter by level here.
                used_created_at_filter = True
                cur.execute(
                    '''
                    SELECT id
                    FROM daily_tests_bank
                    WHERE created_by = ?
                      AND subject = ?
                      AND active = 1
                      AND created_at >= ?
                    ORDER BY id DESC
                    LIMIT ?
                    ''',
                    (user['id'], arena_subject, start_ts, cnt),
                )
                rows = cur.fetchall()

                if not rows:
                    used_created_at_filter = False
                    # Fallback: if created_at boundary mismatches, retry without it.
                    cur.execute(
                        '''
                        SELECT id
                        FROM daily_tests_bank
                        WHERE created_by = ?
                          AND subject = ?
                          AND active = 1
                        ORDER BY id DESC
                        LIMIT ?
                        ''',
                        (user['id'], arena_subject, cnt),
                    )
                    rows = cur.fetchall()
            else:
                used_created_at_filter = True
                cur.execute(
                    '''
                    SELECT id
                    FROM daily_tests_bank
                    WHERE created_by = ?
                      AND subject = ?
                      AND level = ?
                      AND active = 1
                      AND created_at >= ?
                    ORDER BY id DESC
                    LIMIT ?
                    ''',
                    (user['id'], arena_subject, arena_level, start_ts, cnt),
                )
                rows = cur.fetchall()

                if not rows:
                    used_created_at_filter = False
                    cur.execute(
                        '''
                        SELECT id
                        FROM daily_tests_bank
                        WHERE created_by = ?
                          AND subject = ?
                          AND level = ?
                          AND active = 1
                        ORDER BY id DESC
                        LIMIT ?
                        ''',
                        (user['id'], arena_subject, arena_level, cnt),
                    )
                    rows = cur.fetchall()

            # Query returns newest->oldest; reverse to keep stable order.
            bank_ids = list(reversed([int(r['id']) for r in rows]))
        finally:
            conn.close()

        logger.info(
            "teacher_arena_ai_count selected daily bank topic_mode=%s used_created_at_filter=%s bank_ids_count=%s",
            topic_mode,
            used_created_at_filter,
            len(bank_ids),
        )

        logger.info(
            "teacher_arena_ai_count selected daily bank rows count=%s (first=%s last=%s)",
            len(bank_ids),
            bank_ids[0] if bank_ids else None,
            bank_ids[-1] if bank_ids else None,
        )

        if not bank_ids:
            await message.answer(t(lang, "teacher_group_arena_daily_tests_not_found"))
            ts['step'] = None
            return

        # Arena bankiga ko‘chiramiz
        try:
            arena_bank_ids = copy_daily_tests_bank_rows_to_arena_questions(
                bank_ids=bank_ids,
                created_by=user['id'],
            )
        except Exception as e:
            logger.exception(
                "teacher_arena_ai_count copy to arena_questions_bank failed user_id=%s bank_ids_count=%s",
                user.get("id"),
                len(bank_ids),
            )
            await message.answer(t(lang, "teacher_group_arena_copy_failed_with_error", error=str(e)))
            ts['step'] = None
            return

        logger.info(
            "teacher_arena_ai_count copied to arena_questions_bank count=%s",
            len(arena_bank_ids) if arena_bank_ids else 0,
        )

        if not arena_bank_ids:
            await message.answer(t(lang, "teacher_group_arena_copy_failed"))
            ts['step'] = None
            return

        # Session yaratamiz
        try:
            session_id = create_arena_group_session(
                group_id=gid,
                subject=arena_subject,
                level=arena_level,
                question_count=len(arena_bank_ids),
                bank_ids=arena_bank_ids,
                created_by_teacher_id=user['id'],
            )
        except Exception as e:
            logger.exception(
                "teacher_arena_ai_count create_arena_group_session failed user_id=%s gid=%s",
                user.get("id"),
                gid,
            )
            await message.answer(t(lang, "teacher_group_arena_session_create_failed", error=str(e)))
            ts['step'] = None
            return

        logger.info("teacher_arena_ai_count created session_id=%s", session_id)

        ts['step'] = None
        ts['data'] = {}

        await message.answer(
            t(
                lang,
                "teacher_group_arena_ready",
                session_id=session_id,
                question_count=len(arena_bank_ids),
            ),
            reply_markup=InlineKeyboardMarkup(
                inline_keyboard=[
                    [
                        InlineKeyboardButton(
                            text="📨 Talabalarga yuborish",
                            callback_data=f"teacher_arena_send:{session_id}",
                        )
                    ]
                ]
            ),
        )
        return

    # === AI generator count inputs ===
    if ts.get('step') in ('teacher_ai_vocab_await_count', 'teacher_ai_daily_tests_await_count'):
        # Cancel handling during AI generation flows
        if (message.text or '').strip() == t(lang, 'btn_cancel'):
            ts['step'] = None
            ts['data'] = {}
            await message.answer(t(lang, 'select_from_menu'), reply_markup=teacher_main_keyboard(lang))
            return

        if not (message.text or '').strip().isdigit():
            await message.answer(t(lang, "validation_only_number"))
            return

        count = int((message.text or '').strip())
        if count <= 0 or count > 5000:
            await message.answer(t(lang, "validation_count_range", min=1, max=5000))
            return

        from ai_generator import levels_for_ai_generation

        subject = (ts.get('data') or {}).get('subject')
        level = (ts.get('data') or {}).get('level')
        allowed_levels = set(levels_for_ai_generation(subject or ""))
        if subject not in ('English', 'Russian') or level not in allowed_levels:
            ts['step'] = None
            ts['data'] = {}
            await message.answer(t(lang, "invalid_subject_level_state"))
            return

        try:
            if ts.get('step') == 'teacher_ai_vocab_await_count':
                from ai_generator import generate_vocabulary_and_insert
                progress_msg = await message.answer(
                    t(lang, "ai_generation_progress_pct_detail", pct=0, current=0, total=count)
                )
                remaining = count
                batch_size = 50
                requested = generated = inserted = skipped = 0
                while remaining > 0:
                    cur = min(batch_size, remaining)
                    rep = await generate_vocabulary_and_insert(
                        subject=subject,
                        level=level,
                        count=cur,
                        added_by=user['id'],
                    )
                    requested += int(rep.requested or 0)
                    generated += int(rep.generated or 0)
                    inserted += int(rep.inserted or 0)
                    skipped += int(rep.skipped or 0)
                    remaining -= cur
                    done = count - remaining
                    pct = int((done * 100) / max(1, count))
                    try:
                        await progress_msg.edit_text(
                            t(lang, "ai_generation_progress_pct_detail", pct=pct, current=done, total=count)
                        )
                    except Exception:
                        pass
                report = type("R", (), {"requested": requested, "generated": generated, "inserted": inserted, "skipped": skipped})()
                await message.answer(
                    t(
                        lang,
                        "admin_ai_vocab_done_report",
                        requested=report.requested,
                        generated=report.generated,
                        inserted=report.inserted,
                        skipped=report.skipped,
                    )
                )
            else:
                from ai_generator import generate_daily_tests_and_insert
                progress_msg = await message.answer(
                    t(lang, "ai_generation_progress_pct_detail", pct=0, current=0, total=count)
                )
                remaining = count
                batch_size = 50
                requested = generated = inserted = skipped = 0
                while remaining > 0:
                    cur = min(batch_size, remaining)
                    rep = await generate_daily_tests_and_insert(
                        subject=subject,
                        level=level,
                        count=cur,
                        created_by=user['id'],
                    )
                    requested += int(rep.requested or 0)
                    generated += int(rep.generated or 0)
                    inserted += int(rep.inserted or 0)
                    skipped += int(rep.skipped or 0)
                    remaining -= cur
                    done = count - remaining
                    pct = int((done * 100) / max(1, count))
                    try:
                        await progress_msg.edit_text(
                            t(lang, "ai_generation_progress_pct_detail", pct=pct, current=done, total=count)
                        )
                    except Exception:
                        pass
                report = type("R", (), {"requested": requested, "generated": generated, "inserted": inserted, "skipped": skipped})()
                await message.answer(
                    t(
                        lang,
                        "admin_ai_daily_tests_done_report",
                        requested=report.requested,
                        generated=report.generated,
                        inserted=report.inserted,
                        skipped=report.skipped,
                    )
                )
        except Exception as e:
            await message.answer(t(lang, "ai_generation_error_generic", error=str(e)))
        finally:
            ts['step'] = None
            ts['data'] = {}

        await message.answer(t(lang, 'select_from_menu'), reply_markup=teacher_main_keyboard(lang))
        return

    # Do not force access_enabled for teachers; they remain logged in once activated
    
    txt = message.text
    # Teacher menu buttons
    vocab_menu_btn = t(lang, 'teacher_vocab_io_btn')
    attendance_btn = t(lang, 'teacher_attendance_btn')
    daily_tests_upload_btn = t(lang, 'teacher_daily_tests_upload_btn')
    teacher_dcoin_btn = t(lang, 'teacher_dcoin_leaderboard_btn')
    ai_menu_btn = t(lang, 'teacher_ai_generate_btn')
    temp_share_btn = t(lang, 'teacher_temp_share_btn')
    profile_btn = '👤 ' + t(lang, 'my_profile')
    my_groups_btn = '📚 ' + t(lang, 'menu_my_groups')
    arena_btn = '⚔️ ' + t(lang, 'menu_arena')

    if txt == profile_btn:
        await show_my_profile(message.chat.id, user)
        return

    if txt == my_groups_btn:
        await show_my_groups(message.chat.id, user)
        return

    if txt == arena_btn:
        kb = InlineKeyboardMarkup(
            inline_keyboard=[
                [InlineKeyboardButton(text=t(lang, "teacher_group_arena_enable_btn"), callback_data="teacher_arena_group_start")],
                [InlineKeyboardButton(text=t(lang, "arena_rules_button"), callback_data="teacher_arena_rules")],
            ]
        )
        await message.answer(t(lang, "arena_manage_title"), reply_markup=kb)
        return

    if txt == ai_menu_btn:
        kb = InlineKeyboardMarkup(
            inline_keyboard=[
                [InlineKeyboardButton(text=t(lang, "admin_ai_menu_vocab_generate_btn"), callback_data='teacher_ai_vocab_action_generate')],
                [InlineKeyboardButton(text=t(lang, "admin_ai_menu_daily_tests_generate_btn"), callback_data='teacher_ai_daily_tests_action_generate')],
                [InlineKeyboardButton(text=t(lang, "admin_ai_menu_daily_tests_stock_btn"), callback_data='teacher_ai_daily_tests_stock_view')],
                [InlineKeyboardButton(text=t(lang, "admin_ai_menu_daily_tests_history_btn"), callback_data='teacher_ai_daily_tests_history_view')],
            ]
        )
        await message.answer(t(lang, 'choose'), reply_markup=kb)
        return

    if txt == vocab_menu_btn:
        kb = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text=t(lang, 'vocab_import_btn'), callback_data='teacher_vocab_action_import')],
            [InlineKeyboardButton(text=t(lang, 'vocab_export_btn'), callback_data='teacher_vocab_action_export')],
        ])
        await message.answer(t(lang, 'choose'), reply_markup=kb)
        return
    if txt == attendance_btn:
        await _send_teacher_attendance_menu(message.chat.id, user, page=0)
        return

    if txt == temp_share_btn:
        await _send_tshare_groups_list(message.chat.id, user, page=0)
        return

    if txt == daily_tests_upload_btn:
        if not get_daily_test_upload_permission(user['id']):
            await message.answer(t(lang, 'teacher_daily_tests_permission_denied'))
            return

        subject = (user.get('subject') or '').title()
        if subject not in ('English', 'Russian'):
            await message.answer(t(lang, 'invalid_subject'))
            return

        levels = ['A1', 'A2', 'B1', 'B2'] if subject == 'Russian' else ['A1', 'A2', 'B1', 'B2', 'C1']
        kb = InlineKeyboardMarkup(
            inline_keyboard=[
                [
                    InlineKeyboardButton(
                        text=level_ui_label(lang, subject=subject, code=lvl),
                        callback_data=f"daily_tests_level_{lvl}",
                    )
                ]
                for lvl in levels
            ]
        )
        ts['step'] = 'choose_daily_tests_level'
        ts['data'] = {'subject': subject}
        await message.answer(t(lang, 'teacher_daily_tests_level_prompt'), reply_markup=kb)
        return

    if txt == teacher_dcoin_btn:
        subject = (user.get("subject") or "").strip().title()
        if subject not in ("English", "Russian"):
            await message.answer(t(lang, "invalid_subject"))
            return
        await _teacher_send_dcoin_leaderboard(message, lang, subject, 0)
        return

    await message.answer(t(lang, 'select_from_menu'), reply_markup=teacher_main_keyboard(lang))


async def show_my_groups(chat_id: int, user: dict):
    lang = detect_lang_from_user(user)
    groups = _manageable_groups_for_teacher(user['id'])
    
    if not groups:
        await bot.send_message(chat_id, t(lang, 'no_groups'))
        return
    
    text = f"👥 {t(lang, 'my_groups')}\n\n"
    
    for g in groups:
        # Get group statistics
        group_id = g['id']
        
        # Count students in this group
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) as cnt FROM users WHERE group_id=? AND login_type IN (1,2)", (group_id,))
        student_row = cur.fetchone()
        student_count = student_row['cnt'] if student_row else 0
        conn.close()
        
        # Count D'coin users in this group
        conn = get_conn()
        cur = conn.cursor()
        try:
            cur.execute(
                """
                SELECT COUNT(DISTINCT u.id) as cnt
                FROM users u
                JOIN user_subject_dcoins sd ON sd.user_id=u.id
                WHERE u.group_id=? AND u.login_type IN (1,2) AND COALESCE(sd.balance, 0) > 0
                """,
                (group_id,),
            )
            diamond_row = cur.fetchone()
            diamond_count = diamond_row['cnt'] if diamond_row else 0
        except Exception:
            diamond_count = 0
            logger.exception("Failed to query group D'coin users for group_id=%s", group_id)
        finally:
            conn.close()
        
        text += f"📚 {g['name']} ({format_group_level_display(lang, g.get('level'), subject=g.get('subject'))})\n"
        text += f"   👥 O'quvchilar: {student_count} ta\n"
        text += f"   💎 D'coin foydalanuvchilar: {diamond_count} ta\n\n"
    
    await bot.send_message(chat_id, text)


async def show_my_profile(chat_id: int, user: dict):
    lang = detect_lang_from_user(user)
    
    # Teacher uchun to'g'ri ma'lumotlar
    groups = _manageable_groups_for_teacher(user['id'])
    total_students = len({u["id"] for g in groups for u in (get_group_users(g["id"]) or []) if int(u.get("login_type") or 0) in (1, 2)})

    lines = []
    lines.append("👤 " + t(lang, 'my_profile'))
    lines.append("")
    lines.append(f"👨‍🏫 Ism Familya: {user.get('first_name','-')} {user.get('last_name','-')}")
    lines.append(f"📞 Telefon: {user.get('phone') or '-'}")
    lines.append(f"🆔 Login ID: {user.get('login_id') or '-'}")
    lines.append(f"📚 Fan: {user.get('subject') or '-'}")
    lines.append(f"👥 Guruhlar soni: {len(groups)}")
    lines.append(f"👨‍🎓 Jami o'quvchilar: {total_students} ta")

    # Guruhlar ro'yxati (ixtiyoriy, chiroyli)
    if groups:
        lines.append("\n📋 Guruhlar:")
        for g in groups:
            students_in_group = len(get_group_users(g['id']))
            lines.append(
                f"   • {g.get('name')} ({format_group_level_display(lang, g.get('level'), subject=g.get('subject'))}) — {students_in_group} talaba"
            )

    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text=t(lang, 'survey'), callback_data='profile_feedback_menu')],
        [InlineKeyboardButton(text=t(lang, 'choose_lang'), callback_data='profile_choose_lang')],
        [InlineKeyboardButton(text=t(lang, 'logout'), callback_data='logout_me')],
    ])

    await bot.send_message(chat_id, "\n".join(lines), reply_markup=kb)


async def _send_teacher_attendance_menu(chat_id: int, user: dict, page: int = 0, export_mode: bool = False):
    lang = detect_lang_from_user(user)
    groups = [g for g in _manageable_groups_for_teacher(user["id"]) if _has_lesson_today(g)]
    if not groups:
        await bot.send_message(chat_id, t(lang, 'no_lessons_today'))
        return
    total = len(groups)
    total_pages = max(1, (total + LIST_PAGE_SIZE - 1) // LIST_PAGE_SIZE)
    page = max(0, min(page, total_pages - 1))
    sub = groups[page * LIST_PAGE_SIZE:(page + 1) * LIST_PAGE_SIZE]
    items = _format_group_items(lang, sub, include_teacher=True)
    title_key = 'teacher_attendance_export_pick_title' if export_mode else 'teacher_attendance_today_lessons_title'
    text = t(lang, title_key, items=items)
    today = datetime.now(TASK_TZ).strftime("%Y-%m-%d")

    def _pick_cb(g: dict) -> str:
        if export_mode:
            return f"tattmenu:exp_pick:{g['id']}"
        return f"tattmenu:pick:{g['id']}:{today}"

    extra_rows = []
    if not export_mode:
        extra_rows.append([InlineKeyboardButton(text=t(lang, "teacher_attendance_export_btn"), callback_data=f"tattmenu:exp_page:{page}")])
    kb = _build_numbered_keyboard(groups, page, _pick_cb, "tattmenu:page", extra_rows=extra_rows)
    await bot.send_message(chat_id, text, reply_markup=kb)


async def _send_tshare_groups_list(chat_id: int, user: dict, page: int = 0):
    lang = detect_lang_from_user(user)
    groups = get_groups_by_teacher(user["id"]) or []
    if not groups:
        await bot.send_message(chat_id, t(lang, "groups_not_found"))
        return
    total = len(groups)
    total_pages = max(1, (total + LIST_PAGE_SIZE - 1) // LIST_PAGE_SIZE)
    page = max(0, min(page, total_pages - 1))
    sub = groups[page * LIST_PAGE_SIZE:(page + 1) * LIST_PAGE_SIZE]
    text = t(lang, "teacher_temp_share_title", items=_format_group_items(lang, sub, include_teacher=False))
    kb = _build_numbered_keyboard(
        groups,
        page,
        lambda g: f"tshare:grp:{g['id']}",
        "tshare:grp_page",
        extra_rows=[[InlineKeyboardButton(text=t(lang, "teacher_temp_share_mine_btn"), callback_data="tshare:mine:0")]],
    )
    await bot.send_message(chat_id, text, reply_markup=kb)


async def _send_tshare_teachers_list(chat_id: int, user: dict, group_id: int, page: int = 0):
    lang = detect_lang_from_user(user)
    teachers = [x for x in (get_all_teachers() or []) if x.get("id") != user.get("id")]
    if not teachers:
        await bot.send_message(chat_id, t(lang, "teachers_not_found"))
        return
    total = len(teachers)
    total_pages = max(1, (total + LIST_PAGE_SIZE - 1) // LIST_PAGE_SIZE)
    page = max(0, min(page, total_pages - 1))
    sub = teachers[page * LIST_PAGE_SIZE:(page + 1) * LIST_PAGE_SIZE]
    lines = []
    for i, tr in enumerate(sub, start=1):
        lines.append(f"{i}. {_teacher_name(tr)}")
    text = t(lang, "teacher_temp_share_pick_teacher", items="\n".join(lines))
    kb = _build_numbered_keyboard(
        teachers,
        page,
        lambda tr: f"tshare:tch:{group_id}:{tr['id']}",
        f"tshare:tch_page:{group_id}",
    )
    await bot.send_message(chat_id, text, reply_markup=kb)


async def _send_tshare_lessons_picker(chat_id: int, user: dict, page: int = 0):
    lang = detect_lang_from_user(user)
    ts = get_teacher_state(chat_id)
    lessons = (ts.get("data") or {}).get("upcoming_lessons") or []
    selected = set((ts.get("data") or {}).get("selected_lesson_idx") or [])
    total = len(lessons)
    total_pages = max(1, (total + LIST_PAGE_SIZE - 1) // LIST_PAGE_SIZE)
    page = max(0, min(page, total_pages - 1))
    sub = lessons[page * LIST_PAGE_SIZE:(page + 1) * LIST_PAGE_SIZE]
    lines = []
    for i, ls in enumerate(sub, start=1):
        real_idx = page * LIST_PAGE_SIZE + i - 1
        mark = "✅" if real_idx in selected else "▫️"
        lines.append(
            f"{i}. {mark} {_fmt_date(ls.get('lesson_date'))} ({ls.get('weekday')}) {ls.get('lesson_start')}-{ls.get('lesson_end')}"
        )
    text = t(lang, "teacher_temp_share_lessons_pick_title", selected=len(selected), needed=len(lessons), items="\n".join(lines))

    def _pick(ls: dict) -> str:
        idx = lessons.index(ls)
        return f"tshare:lsn:{idx}:{page}"

    kb = _build_numbered_keyboard(
        lessons,
        page,
        _pick,
        "tshare:lsn_page",
        extra_rows=[[
            InlineKeyboardButton(text=t(lang, "teacher_temp_share_confirm_btn"), callback_data="tshare:confirm"),
            InlineKeyboardButton(text=t(lang, "teacher_temp_share_deny_btn"), callback_data="tshare:deny"),
        ]],
    )
    prev_mid = int((ts.get("data") or {}).get("picker_message_id") or 0)
    if prev_mid:
        try:
            await bot.delete_message(chat_id=chat_id, message_id=prev_mid)
        except Exception:
            pass
    msg = await bot.send_message(chat_id, text, reply_markup=kb)
    ts["data"]["picker_message_id"] = msg.message_id

async def show_teacher_feedback_menu(chat_id: int, user: dict, lang: str):
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text=t(lang, 'feedback_btn_anonymous'), callback_data="t_feedback_anonymous")],
        [InlineKeyboardButton(text=t(lang, 'feedback_btn_named'), callback_data="t_feedback_named")],
    ])
    await bot.send_message(
        chat_id,
        t(lang, 'teacher_feedback_menu_title'),
        reply_markup=kb,
        parse_mode="HTML",
    )


async def handle_teacher_feedback_text(message: Message, user: dict, lang: str):
    """O‘qituvchi fikr matni — avval tasdiqlash."""
    ts = get_teacher_state(message.chat.id)
    if ts.get('step') != 'feedback_wait':
        return
    text = (message.text or "").strip()
    if not text:
        await message.answer(t(lang, 'feedback_empty_error'))
        return
    is_anonymous = (ts.get('data') or {}).get('anonymous', True)
    ts['step'] = 'feedback_confirm'
    ts['data']['draft'] = text
    ts['data']['anonymous'] = is_anonymous
    safe = html_module.escape(text[:3800])
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text=t(lang, 'feedback_btn_send_admin'), callback_data="t_fb_ok")],
        [InlineKeyboardButton(text=t(lang, 'feedback_btn_cancel'), callback_data="t_fb_cancel")],
    ])
    await message.answer(
        t(lang, 'feedback_confirm_prompt', draft=safe),
        reply_markup=kb,
        parse_mode="HTML",
    )


async def _notify_admins_teacher_message(html_body: str, sender_tg_id: str | None = None):
    global admin_bot
    b = admin_bot
    if not b and ADMIN_BOT_TOKEN:
        b = Bot(token=ADMIN_BOT_TOKEN, default=DefaultBotProperties(parse_mode="HTML"))
    if not b or not ADMIN_CHAT_IDS:
        return
    for aid in ADMIN_CHAT_IDS:
        try:
            reply_markup = None
            if sender_tg_id and str(sender_tg_id).strip().isdigit():
                reply_markup = InlineKeyboardMarkup(inline_keyboard=[[
                    InlineKeyboardButton(
                        text="✉️ Javob yozish",
                        url=f"tg://user?id={int(sender_tg_id)}",
                    )
                ]])
            await b.send_message(aid, html_body, parse_mode="HTML", reply_markup=reply_markup)
        except Exception as e:
            logger.exception("Teacher xabari admin ga yuborilmadi: %s", e)


def _teacher_feedback_admin_profile(user: dict) -> str:
    groups = _manageable_groups_for_teacher(user['id']) or []
    group_names = ", ".join([(g.get('name') or '-') for g in groups]) if groups else '-'
    subjects_text = (user.get('subject') or '-')
    dcoin = get_dcoins(user['id'])
    total_students = get_teacher_total_students(user['id'])
    return (
        f"👤 Ism: <b>{html_module.escape((user.get('first_name') or '-'))}</b>\n"
        f"👤 Familya: <b>{html_module.escape((user.get('last_name') or '-'))}</b>\n"
        f"📞 Telefon: <b>{html_module.escape((user.get('phone') or '-'))}</b>\n"
        f"📚 Fan(lar): <b>{html_module.escape(subjects_text)}</b>\n"
        f"👥 Guruh(lar): <b>{html_module.escape(group_names)}</b>\n"
        f"👨‍🎓 Jami o'quvchi: <b>{total_students}</b>\n"
        f"💎 D'coin: <b>{dcoin:.1f}</b>\n"
    )


@dp.callback_query(lambda c: c.data == "profile_feedback_menu")
async def handle_teacher_profile_feedback_menu(callback: CallbackQuery):
    user = get_user_by_telegram(str(callback.from_user.id))
    lang = detect_lang_from_user(user or callback.from_user)
    if not user or user.get("login_type") != 3:
        await callback.answer()
        await callback.message.answer(t(lang, 'please_send_start'))
        return
    await show_teacher_feedback_menu(callback.message.chat.id, user, lang)
    await callback.answer()


@dp.callback_query(lambda c: c.data == "profile_choose_lang")
async def handle_teacher_profile_choose_lang(callback: CallbackQuery):
    user = get_user_by_telegram(str(callback.from_user.id))
    lang = detect_lang_from_user(user or callback.from_user)
    if not user or user.get("login_type") != 3:
        await callback.answer()
        await callback.message.answer(t(lang, 'please_send_start'))
        return
    await callback.message.answer(t(lang, 'choose_lang'), reply_markup=create_language_selection_keyboard_for_self(lang))
    await callback.answer()


@dp.callback_query(lambda c: c.data in ("t_feedback_anonymous", "t_feedback_named"))
async def handle_teacher_feedback_mode(callback: CallbackQuery):
    user = get_user_by_telegram(str(callback.from_user.id))
    lang = detect_lang_from_user(user or callback.from_user)
    if not user or user.get("login_type") != 3:
        await callback.answer()
        await callback.message.answer(t(lang, 'please_send_start'))
        return
    ts = get_teacher_state(callback.message.chat.id)
    ts['step'] = 'feedback_wait'
    ts['data'] = {'anonymous': callback.data.endswith("anonymous")}
    if callback.data.endswith("anonymous"):
        prompt = t(lang, 'feedback_prompt_anonymous')
    else:
        prompt = t(
            lang,
            'feedback_prompt_named',
            author=f"{user.get('first_name', '')} {user.get('last_name', '')}",
        )
    await callback.message.answer(prompt, parse_mode="HTML")
    await callback.answer()


@dp.callback_query(lambda c: c.data in ("t_fb_ok", "t_fb_cancel"))
async def handle_teacher_feedback_confirm(callback: CallbackQuery):
    user = get_user_by_telegram(str(callback.from_user.id))
    if not user or user.get("login_type") != 3:
        await callback.answer()
        return
    ts = get_teacher_state(callback.message.chat.id)
    if ts.get('step') != 'feedback_confirm':
        await callback.answer(t(detect_lang_from_user(user or callback.from_user), 'feedback_action_expired'), show_alert=True)
        return
    if callback.data == "t_fb_cancel":
        ts['step'] = None
        ts['data'] = {}
        lang = detect_lang_from_user(user or callback.from_user)
        try:
            await callback.message.edit_text(t(lang, 'feedback_send_cancelled'))
        except Exception:
            await callback.message.answer(t(lang, 'feedback_send_cancelled'))
        await callback.answer()
        return

    draft = (ts.get('data') or {}).get('draft', "")
    is_anonymous = (ts.get('data') or {}).get('anonymous', True)
    add_feedback(user['id'], draft, is_anonymous)
    lang = detect_lang_from_user(user or callback.from_user)
    who = "Anonim o‘qituvchi" if is_anonymous else "O‘qituvchi (ism bilan yuborilgan)"
    profile = _teacher_feedback_admin_profile(user)
    admin_html = (
        "💬 <b>O‘qituvchi xabari</b>\n"
        f"👤 {who}\n"
        f"🆔 user_id: <code>{user['id']}</code>\n"
        f"🆔 login_id: <code>{html_module.escape(user.get('login_id') or '-')}</code>\n"
        f"📱 tg: <code>{user.get('telegram_id') or '—'}</code>\n\n"
        f"{profile}\n"
        "📝 <b>Xabar matni:</b>\n"
        f"<pre>{html_module.escape(draft[:3500])}</pre>"
    )
    await _notify_admins_teacher_message(admin_html, user.get('telegram_id'))
    ts['step'] = None
    ts['data'] = {}
    try:
        await callback.message.edit_text(t(lang, 'feedback_sent_success'), parse_mode="HTML")
    except Exception:
        await callback.message.answer(t(lang, 'feedback_sent_success'), parse_mode="HTML")
    await callback.answer()


@dp.callback_query(lambda c: c.data == 'logout_me')
async def handle_logout_me_teacher(callback: CallbackQuery):
    from db import logout_user_by_telegram
    user = get_user_by_telegram(str(callback.from_user.id))
    lang = detect_lang_from_user(user or callback.from_user)
    logout_user_by_telegram(str(callback.from_user.id))
    clear_login_state(str(callback.from_user.id))
    reset_teacher_state(callback.message.chat.id)
    await callback.answer()
    await callback.message.answer(t(lang, 'logged_out'), reply_markup=ReplyKeyboardRemove())


@dp.callback_query(lambda c: c.data.startswith("tattmenu:"))
async def handle_teacher_attendance_menu_callbacks(callback: CallbackQuery):
    user = get_user_by_telegram(str(callback.from_user.id))
    lang = detect_lang_from_user(user or callback.from_user)
    if not user or user.get("login_type") != 3:
        await callback.answer()
        return
    data = callback.data or ""
    parts = data.split(":")
    action = parts[1] if len(parts) > 1 else ""
    await callback.answer()
    if action == "page":
        await _send_teacher_attendance_menu(callback.message.chat.id, user, page=int(parts[2]), export_mode=False)
    elif action == "exp_page":
        await _send_teacher_attendance_menu(callback.message.chat.id, user, page=int(parts[2]), export_mode=True)
    elif action == "pick":
        gid = int(parts[2])
        date_str = parts[3]
        group = get_group(gid)
        if not _teacher_can_manage_group(user, group):
            return
        from attendance_manager import ensure_session, send_attendance_panel
        session = ensure_session(gid, date_str)
        await send_attendance_panel(
            bot,
            callback.message.chat.id,
            session["id"],
            gid,
            date_str,
            0,
            panel_owner="teacher",
        )
    elif action == "exp_pick":
        gid = int(parts[2])
        group = get_group(gid)
        if not _teacher_can_manage_group(user, group):
            return
        from vocabulary import export_group_attendance_to_xlsx
        bio, fname = export_group_attendance_to_xlsx(gid)
        gname = group.get("name") or t(lang, "attendance_fallback_group_name")
        await callback.message.answer_document(
            BufferedInputFile(bio.getvalue(), filename=fname),
            caption=t(lang, "teacher_attendance_export_caption", group_name=gname),
        )


@dp.callback_query(lambda c: c.data.startswith("tshare:"))
async def handle_teacher_temp_share_callbacks(callback: CallbackQuery):
    user = get_user_by_telegram(str(callback.from_user.id))
    lang = detect_lang_from_user(user or callback.from_user)
    if not user or user.get("login_type") != 3:
        await callback.answer()
        return
    data = callback.data or ""
    parts = data.split(":")
    action = parts[1] if len(parts) > 1 else ""
    ts = get_teacher_state(callback.message.chat.id)
    await callback.answer()

    if action == "grp_page":
        await _send_tshare_groups_list(callback.message.chat.id, user, page=int(parts[2]))
        return
    if action == "grp":
        gid = int(parts[2])
        group = get_group(gid)
        if not group or group.get("teacher_id") != user.get("id"):
            return
        ts["step"] = "tshare_pick_teacher"
        ts["data"] = {"group_id": gid}
        await _send_tshare_teachers_list(callback.message.chat.id, user, gid, page=0)
        return
    if action == "tch_page":
        gid = int(parts[2])
        page = int(parts[3])
        await _send_tshare_teachers_list(callback.message.chat.id, user, gid, page=page)
        return
    if action == "tch":
        gid = int(parts[2])
        tid = int(parts[3])
        if tid == user.get("id"):
            await callback.message.answer(t(lang, "err_invalid_choice"))
            return
        group = get_group(gid)
        if not group:
            await callback.message.answer(t(lang, "group_not_found"))
            return
        upcoming = _compute_upcoming_lessons(group, take=20)
        if not upcoming:
            await callback.message.answer(t(lang, "teacher_temp_share_no_upcoming"))
            return
        ts["step"] = "tshare_pick_lessons"
        ts["data"] = {
            "group_id": gid,
            "temp_teacher_id": tid,
            "upcoming_lessons": upcoming,
            "selected_lesson_idx": [],
            "picker_message_id": 0,
        }
        await _send_tshare_lessons_picker(callback.message.chat.id, user, page=0)
        return
    if action == "lsn":
        idx = int(parts[2])
        page = int(parts[3])
        selected = set((ts.get("data") or {}).get("selected_lesson_idx") or [])
        if idx in selected:
            selected.remove(idx)
        else:
            selected.add(idx)
        ts["data"]["selected_lesson_idx"] = sorted(selected)
        await _send_tshare_lessons_picker(callback.message.chat.id, user, page=page)
        return
    if action == "lsn_page":
        await _send_tshare_lessons_picker(callback.message.chat.id, user, page=int(parts[2]))
        return
    if action == "deny":
        prev_mid = int((ts.get("data") or {}).get("picker_message_id") or 0)
        if prev_mid:
            try:
                await bot.delete_message(chat_id=callback.message.chat.id, message_id=prev_mid)
            except Exception:
                pass
        ts["step"] = None
        ts["data"] = {}
        await callback.message.answer(t(lang, "operation_cancelled"))
        return
    if action == "confirm":
        selected = (ts.get("data") or {}).get("selected_lesson_idx") or []
        if len(selected) < 1:
            await callback.message.answer(t(lang, "teacher_temp_share_pick_required"))
            return
        gid = int((ts.get("data") or {}).get("group_id") or 0)
        tid = int((ts.get("data") or {}).get("temp_teacher_id") or 0)
        group = get_group(gid)
        teacher = get_user_by_id(tid)
        lessons = (ts.get("data") or {}).get("upcoming_lessons") or []
        picked = []
        for idx in selected:
            if 0 <= idx < len(lessons):
                ls = lessons[idx]
                picked.append(f"- {_fmt_date(ls.get('lesson_date'))} ({ls.get('weekday')}) {ls.get('lesson_start')}-{ls.get('lesson_end')}")
        ts["step"] = "tshare_confirm"
        kb = InlineKeyboardMarkup(inline_keyboard=[[
            InlineKeyboardButton(text=t(lang, "teacher_temp_share_confirm_btn"), callback_data="tshare:do_confirm"),
            InlineKeyboardButton(text=t(lang, "teacher_temp_share_deny_btn"), callback_data="tshare:deny"),
        ]])
        await callback.message.answer(
            t(lang, "teacher_temp_share_confirm_text", group=group.get("name", "-"), teacher=_teacher_name(teacher or {}), count=len(selected)) + "\n\n" + "\n".join(picked),
            reply_markup=kb,
        )
        return
    if action == "do_confirm":
        datax = ts.get("data") or {}
        gid = int(datax.get("group_id") or 0)
        tid = int(datax.get("temp_teacher_id") or 0)
        group = get_group(gid)
        if not group or group.get("teacher_id") != user.get("id"):
            return
        lessons = datax.get("upcoming_lessons") or []
        selected = datax.get("selected_lesson_idx") or []
        chosen_lessons = []
        for idx in selected:
            if 0 <= idx < len(lessons):
                ls = lessons[idx]
                chosen_lessons.append(ls)
                create_temporary_group_assignment(
                    group_id=gid,
                    owner_teacher_id=user["id"],
                    temp_teacher_id=tid,
                    lesson_date=ls.get("lesson_date"),
                    lesson_start=ls.get("lesson_start"),
                    lesson_end=ls.get("lesson_end"),
                )
        temp_teacher = get_user_by_id(tid) or {}
        owner_name = _teacher_name(user)
        group_name = group.get("name", "-")
        details_owner = _temp_assignment_meta(group, lang, chosen_lessons)
        owner_msg = t(lang, "teacher_temp_share_notify_owner", group=group_name, temp_teacher=_teacher_name(temp_teacher), count=len(selected))
        await callback.message.answer(f"{owner_msg}\n\n{details_owner}")
        try:
            if temp_teacher.get("telegram_id"):
                details_temp = _temp_assignment_meta(group, detect_lang_from_user(temp_teacher), chosen_lessons)
                await bot.send_message(
                    int(temp_teacher["telegram_id"]),
                    t(
                        detect_lang_from_user(temp_teacher),
                        "teacher_temp_share_notify_temp",
                        group=group_name,
                        owner_teacher=owner_name,
                        count=len(selected),
                    )
                    + f"\n\n{details_temp}",
                )
        except Exception:
            logger.exception("Failed to notify temporary teacher")
        targets = _temp_assignment_admin_chat_targets(group)
        admin_lang = detect_lang_from_user(user)
        details_admin = _temp_assignment_meta(group, admin_lang, chosen_lessons)
        for tg_id in targets:
            try:
                await bot.send_message(
                    tg_id,
                    t(admin_lang, "teacher_temp_share_notify_admin", group=group_name, owner_teacher=owner_name, temp_teacher=_teacher_name(temp_teacher), count=len(selected))
                    + f"\n\n{details_admin}",
                )
            except Exception:
                logger.exception("Failed to notify admin target tg_id=%s for temporary assignment", tg_id)
        prev_mid = int((ts.get("data") or {}).get("picker_message_id") or 0)
        if prev_mid:
            try:
                await bot.delete_message(chat_id=callback.message.chat.id, message_id=prev_mid)
            except Exception:
                pass
        ts["step"] = None
        ts["data"] = {}
        return
    if action == "mine":
        rows = get_active_temporary_assignments_by_owner(user["id"]) or []
        if not rows:
            await callback.message.answer(t(lang, "no_data_found"))
            return
        grouped = {}
        for r in rows:
            key = (r["group_id"], r["temp_teacher_id"])
            grouped.setdefault(key, {"group_id": r["group_id"], "group_name": r.get("group_name"), "group_level": r.get("group_level"), "temp_teacher_id": r["temp_teacher_id"], "count": 0})
            grouped[key]["count"] += 1
        items = list(grouped.values())
        page = int(parts[2])
        total = len(items)
        total_pages = max(1, (total + LIST_PAGE_SIZE - 1) // LIST_PAGE_SIZE)
        page = max(0, min(page, total_pages - 1))
        sub = items[page * LIST_PAGE_SIZE:(page + 1) * LIST_PAGE_SIZE]
        lines = []
        for i, it in enumerate(sub, start=1):
            temp_teacher = get_user_by_id(it["temp_teacher_id"]) or {}
            students = len(get_group_users(it["group_id"]))
            lines.append(f"{i}. Group name: {it.get('group_name')}")
            lines.append(f"   Vaqtinchalik biriktirilgan teacher: {_teacher_name(temp_teacher)}")
            lines.append(f"   Group level: {it.get('group_level') or '-'}")
            lines.append(f"   Students in group: {students}")
        text = t(lang, "teacher_temp_share_my_active_title", items="\n".join(lines))
        kb = _build_numbered_keyboard(
            items,
            page,
            lambda it: f"tshare:mine_pick:{it['group_id']}:{it['temp_teacher_id']}",
            "tshare:mine",
        )
        await callback.message.answer(text, reply_markup=kb)
        return
    if action == "mine_pick":
        gid = int(parts[2])
        tid = int(parts[3])
        group = get_group(gid)
        teacher = get_user_by_id(tid) or {}
        kb = InlineKeyboardMarkup(inline_keyboard=[[
            InlineKeyboardButton(text=t(lang, "teacher_temp_share_cancel_btn"), callback_data=f"tshare:cancel:{gid}:{tid}")
        ]])
        await callback.message.answer(
            t(lang, "teacher_temp_share_cancel_ask", group=(group or {}).get("name", "-"), teacher=_teacher_name(teacher)),
            reply_markup=kb,
        )
        return
    if action == "cancel":
        gid = int(parts[2])
        tid = int(parts[3])
        active_rows = get_active_temporary_assignments_for_pair(user["id"], gid, tid)
        cancel_temporary_assignments_for_pair(user["id"], gid, tid)
        temp_teacher = get_user_by_id(tid) or {}
        group = get_group(gid) or {}
        chosen_lessons = []
        for r in active_rows:
            chosen_lessons.append(
                {
                    "lesson_date": r.get("lesson_date"),
                    "lesson_start": r.get("lesson_start") or "-",
                    "lesson_end": r.get("lesson_end") or "-",
                    "weekday": datetime.strptime(r.get("lesson_date"), "%Y-%m-%d").strftime("%A")
                    if r.get("lesson_date")
                    else "-",
                }
            )
        details_owner = _temp_assignment_meta(group, lang, chosen_lessons)
        owner_name = _teacher_name(user)
        try:
            if temp_teacher.get("telegram_id"):
                t_lang = detect_lang_from_user(temp_teacher)
                details_temp = _temp_assignment_meta(group, t_lang, chosen_lessons)
                await bot.send_message(
                    int(temp_teacher["telegram_id"]),
                    t(t_lang, "teacher_temp_share_notify_temp_cancel", group=group.get("name", "-"), owner_teacher=owner_name, count=len(chosen_lessons))
                    + f"\n\n{details_temp}",
                )
        except Exception:
            logger.exception("Failed to notify temp teacher cancellation")
        try:
            await callback.message.answer(
                t(lang, "teacher_temp_share_notify_owner_cancel", group=group.get("name", "-"), temp_teacher=_teacher_name(temp_teacher), count=len(chosen_lessons))
                + f"\n\n{details_owner}"
            )
        except Exception:
            logger.exception("Failed to notify owner teacher cancellation")
        targets = _temp_assignment_admin_chat_targets(group)
        for tg_id in targets:
            try:
                await bot.send_message(
                    tg_id,
                    t(
                        detect_lang_from_user(user),
                        "teacher_temp_share_notify_admin_cancel",
                        group=group.get("name", "-"),
                        owner_teacher=owner_name,
                        temp_teacher=_teacher_name(temp_teacher),
                        count=len(chosen_lessons),
                    )
                    + f"\n\n{details_owner}",
                )
            except Exception:
                logger.exception("Failed to notify admin target tg_id=%s for temporary assignment cancellation", tg_id)
        await callback.message.answer(t(lang, "teacher_temp_share_cancel_done"))
        return


@dp.callback_query(lambda c: c.data.startswith("att_open_"))
async def handle_teacher_att_open(callback: CallbackQuery):
    user = get_user_by_telegram(str(callback.from_user.id))
    lang = detect_lang_from_user(user or callback.from_user)
    if not user or user.get("login_type") != 3:
        await callback.answer()
        await callback.message.answer(t(lang, 'please_send_start'))
        return
    parts = callback.data.split("_")
    group_id = int(parts[2])
    date_str = parts[3]
    if is_lesson_holiday(date_str) or is_lesson_otmen_date_cancelled(date_str):
        await callback.answer(t(lang, "lesson_canceled_for_date_alert", date=date_str), show_alert=True)
        return
    group = get_group(group_id)
    if not _teacher_can_manage_group(user, group):
        await callback.answer()
        return
    from attendance_manager import ensure_session, send_attendance_panel
    session = ensure_session(group_id, date_str)
    await callback.answer()
    await send_attendance_panel(
        bot,
        callback.message.chat.id,
        session["id"],
        group_id,
        date_str,
        0,
        panel_owner="teacher",
    )


@dp.callback_query(lambda c: c.data == "att_noop" or c.data.startswith("att_page_") or c.data.startswith("att_mark_") or c.data.startswith("att_finish_"))
async def handle_teacher_att_actions(callback: CallbackQuery):
    data = callback.data
    user = get_user_by_telegram(str(callback.from_user.id))
    if data == "att_noop":
        await callback.answer()
        return
    if not user or user.get("login_type") != 3:
        await callback.answer()
        return
    if data.startswith("att_page_"):
        parts = data.split("_")
        session_id = int(parts[2])
        page = int(parts[3])
        session = get_session(session_id)
        if not session:
            await callback.answer()
            return
        group = get_group(session["group_id"])
        if not _teacher_can_manage_group(user, group):
            await callback.answer()
            return
        tlang = detect_lang_from_user(user or callback.from_user)
        teacher_kb = build_attendance_keyboard(
            session_id,
            session["group_id"],
            session["date"],
            page,
            lang=get_panel_ui_lang(session_id, "teacher", tlang),
        )
        admin_kb = build_attendance_keyboard(
            session_id,
            session["group_id"],
            session["date"],
            page,
            lang=get_panel_ui_lang(session_id, "admin", "uz"),
        )
        # Telegram sometimes throws "message is not modified" when the markup is unchanged.
        try:
            await callback.message.edit_reply_markup(reply_markup=teacher_kb)
        except Exception as e:
            if "message is not modified" not in str(e).lower():
                raise
        # Realtime sync: update admin's panel too (if available).
        try:
            from attendance_manager import _get_panel_message
            admin_panel = _get_panel_message(session_id, "admin")
            if admin_panel:
                admin_chat_id, admin_message_id = admin_panel
                import admin_bot
                if getattr(admin_bot, "bot", None):
                    await admin_bot.bot.edit_message_reply_markup(
                        chat_id=admin_chat_id,
                        message_id=admin_message_id,
                        reply_markup=admin_kb,
                    )
        except Exception:
            logger.exception("Failed to sync admin attendance panel on page nav")
        await callback.answer()
        return
    if data.startswith("att_mark_"):
        parts = data.split("_")
        session_id = int(parts[2])
        student_id = int(parts[3])
        status = parts[4]
        page = int(parts[5])
        session = get_session(session_id)
        if not session or session.get("status") == "closed":
            await callback.answer()
            return
        group = get_group(session["group_id"])
        if not _teacher_can_manage_group(user, group):
            await callback.answer()
            return
        add_attendance(student_id, session["group_id"], session["date"], status=status)
        tlang = detect_lang_from_user(user or callback.from_user)
        teacher_kb = build_attendance_keyboard(
            session_id,
            session["group_id"],
            session["date"],
            page,
            lang=get_panel_ui_lang(session_id, "teacher", tlang),
        )
        admin_kb = build_attendance_keyboard(
            session_id,
            session["group_id"],
            session["date"],
            page,
            lang=get_panel_ui_lang(session_id, "admin", "uz"),
        )
        try:
            await callback.message.edit_reply_markup(reply_markup=teacher_kb)
        except Exception as e:
            if "message is not modified" not in str(e).lower():
                raise
        # Realtime sync: update admin's panel too (if available).
        try:
            from attendance_manager import _get_panel_message
            admin_panel = _get_panel_message(session_id, "admin")
            if admin_panel:
                admin_chat_id, admin_message_id = admin_panel
                import admin_bot
                if getattr(admin_bot, "bot", None):
                    await admin_bot.bot.edit_message_reply_markup(
                        chat_id=admin_chat_id,
                        message_id=admin_message_id,
                        reply_markup=admin_kb,
                    )
        except Exception:
            logger.exception("Failed to sync admin attendance panel on mark")
        await callback.answer()
        return
    if data.startswith("att_finish_"):
        session_id = int(data.split("_")[-1])
        session = get_session(session_id)
        if not session:
            await callback.answer()
            return
        group = get_group(session["group_id"])
        if not _teacher_can_manage_group(user, group):
            await callback.answer()
            return
        set_session_closed(session_id)
        try:
            from attendance_manager import finalize_attendance_session
            await finalize_attendance_session(
                bot,
                session_id,
                admin_chat_ids=ADMIN_CHAT_IDS,
                student_notify_bot=student_notify_bot,
            )
        except Exception:
            logger.exception("Failed to finalize attendance session from teacher finish callback")
        await callback.answer()
        await callback.message.answer(t(detect_lang_from_user(user or callback.from_user), 'attendance_done_closed'))
        # notify admins via teacher bot (best-effort: cannot access admin IDs list here reliably)
        return


@dp.callback_query(lambda c: c.data in ('teacher_vocab_action_import', 'teacher_vocab_action_export'))
async def handle_teacher_vocab_action(callback: CallbackQuery):
    action = callback.data.split('_')[-1]  # import/export
    user = get_user_by_telegram(str(callback.from_user.id))
    lang = detect_lang_from_user(user or callback.from_user)
    if not user or user.get('login_type') != 3:
        await callback.answer()
        await callback.message.answer(t(lang, 'please_send_start'))
        return

    subj = (user.get('subject') or '').title()
    state = get_teacher_state(callback.message.chat.id)

    if action == 'import':
        logger.info(f"📥 Teacher vocab import action user_id={callback.from_user.id}")
        # Directly use teacher's subject since "Both" is removed
        state['step'] = 'await_vocab_file'
        state['data']['subject'] = subj
        await callback.message.answer(t(lang, 'send_vocab_file'), reply_markup=cancel_keyboard(lang))
        
        # ✅ Yangi to'g'ri template (user提供的实际文件)
        try:
            # Load the actual example file based on subject
            if subj.lower() == 'english':
                example_file = 'Vocabulary_importing_list_for_english.xlsx'
            else:
                example_file = 'Vocabulary_importing_list_for_russian.xlsx'
            
            # Read the actual template file
            with open(example_file, 'rb') as f:
                file_data = f.read()
            
            # Send the actual template file
            await bot.send_document(
                callback.message.chat.id,
                BufferedInputFile(file_data, filename="Vocabulary_importing_list_template.xlsx"),
                caption=t(lang, 'teacher_vocab_template_caption_full', subject=subj),
                parse_mode='HTML'
            )
                
        except FileNotFoundError:
            # Fallback to generated template if file not found
            await callback.message.answer(t(lang, 'teacher_template_not_found_generating'))
            
            # Generate fallback template
            from openpyxl import Workbook
            from openpyxl.styles import Font
            import io

            wb = Workbook()
            ws = wb.active
            ws.title = "Template"
            
            # Headers based on subject
            if subj.lower() == 'english':
                headers = ['Level', 'Word', 'translation_uz', 'translation_ru', 'Definition', 'Example Sentence 1', 'Example Sentence 2']
            else:  # Russian
                headers = ['Level', 'Word', 'translation_uz', 'Definition', 'Example Sentence 1', 'Example Sentence 2']
            
            ws.append(headers)
            
            # Make headers bold
            for col in range(1, len(headers) + 1):
                ws.cell(row=1, column=col).font = Font(bold=True)
            
            bio = io.BytesIO()
            wb.save(bio)
            bio.seek(0)
            
            await bot.send_document(
                callback.message.chat.id,
                BufferedInputFile(bio.getvalue(), filename="Vocabulary_importing_list_template.xlsx"),
                caption=t(lang, 'teacher_vocab_template_caption_short', subject=subj),
                parse_mode='HTML'
            )
            
        except Exception as e:
            logger.error(f"Template yuborish xatosi: {e}")
            await callback.message.answer(t(lang, 'teacher_template_send_error'))
        except Exception as e:
            logger.error(f"Template yuborish xatosi: {e}")
        await callback.answer()
        return


@dp.callback_query(lambda c: c.data == 'teacher_ai_vocab_action_generate')
async def handle_teacher_ai_vocab_start(callback: CallbackQuery):
    user = get_user_by_telegram(str(callback.from_user.id))
    lang = detect_lang_from_user(user or callback.from_user)
    if not user or user.get('login_type') != 3:
        await callback.answer()
        await callback.message.answer(t(lang, 'please_send_start'))
        return

    ts = get_teacher_state(callback.message.chat.id)
    subject = (user.get('subject') or '').title()
    ts['data'] = ts.get('data') or {}
    ts['data']['subject'] = subject
    ts['step'] = 'teacher_ai_vocab_choose_level'

    from ai_generator import levels_for_ai_generation

    lvls = levels_for_ai_generation(subject)
    kb = InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton(
                    text=level_ui_label(lang, subject=subject, code=lvl),
                    callback_data=f"teacher_ai_vocab_level_{lvl}",
                )
            ]
            for lvl in lvls
        ]
    )
    await callback.message.answer(t(lang, 'teacher_daily_tests_level_prompt'), reply_markup=kb)
    await callback.answer()


@dp.callback_query(lambda c: c.data.startswith('teacher_ai_vocab_level_'))
async def handle_teacher_ai_vocab_level_pick(callback: CallbackQuery):
    user = get_user_by_telegram(str(callback.from_user.id))
    lang = detect_lang_from_user(user or callback.from_user)
    if not user or user.get('login_type') != 3:
        await callback.answer()
        return

    ts = get_teacher_state(callback.message.chat.id)
    level = callback.data.split('teacher_ai_vocab_level_', 1)[-1].strip().upper()
    from ai_generator import levels_for_ai_generation

    subj = (ts.get('data') or {}).get('subject') or ''
    if level not in set(levels_for_ai_generation(subj)):
        await callback.answer(t(lang, 'invalid_subject_level_state'), show_alert=True)
        return
    ts['data'] = ts.get('data') or {}
    ts['data']['level'] = level
    ts['step'] = 'teacher_ai_vocab_await_count'

    await callback.answer()
    lvl_lbl = level_ui_label(lang, subject=subj, code=level)
    await callback.message.answer(
        f"🤖 {ts['data'].get('subject')} uchun {lvl_lbl} levelga nechta vocabulary generatsiya qilinsin?\n(son kiriting, masalan: 50)",
        reply_markup=cancel_keyboard(lang),
    )


@dp.callback_query(lambda c: c.data == 'teacher_ai_daily_tests_action_generate')
async def handle_teacher_ai_daily_tests_start(callback: CallbackQuery):
    user = get_user_by_telegram(str(callback.from_user.id))
    lang = detect_lang_from_user(user or callback.from_user)
    if not user or user.get('login_type') != 3:
        await callback.answer()
        await callback.message.answer(t(lang, 'please_send_start'))
        return

    ts = get_teacher_state(callback.message.chat.id)
    subject = (user.get('subject') or '').title()
    ts['data'] = ts.get('data') or {}
    ts['data']['subject'] = subject
    ts['step'] = 'teacher_ai_daily_tests_choose_level'

    from ai_generator import levels_for_ai_generation

    lvls = levels_for_ai_generation(subject)
    kb = InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton(
                    text=level_ui_label(lang, subject=subject, code=lvl),
                    callback_data=f"teacher_ai_daily_tests_level_{lvl}",
                )
            ]
            for lvl in lvls
        ]
    )
    await callback.message.answer(t(lang, 'teacher_daily_tests_level_prompt'), reply_markup=kb)
    await callback.answer()


@dp.callback_query(lambda c: c.data.startswith('teacher_ai_daily_tests_level_'))
async def handle_teacher_ai_daily_tests_level_pick(callback: CallbackQuery):
    user = get_user_by_telegram(str(callback.from_user.id))
    lang = detect_lang_from_user(user or callback.from_user)
    if not user or user.get('login_type') != 3:
        await callback.answer()
        return

    ts = get_teacher_state(callback.message.chat.id)
    level = callback.data.split('teacher_ai_daily_tests_level_', 1)[-1].strip().upper()
    from ai_generator import levels_for_ai_generation

    subj = (ts.get('data') or {}).get('subject') or ''
    if level not in set(levels_for_ai_generation(subj)):
        await callback.answer(t(lang, 'invalid_subject_level_state'), show_alert=True)
        return
    ts['data'] = ts.get('data') or {}
    ts['data']['level'] = level
    ts['step'] = 'teacher_ai_daily_tests_await_count'

    await callback.answer()
    lvl_lbl = level_ui_label(lang, subject=subj, code=level)
    await callback.message.answer(
        f"📅 {ts['data'].get('subject')} uchun {lvl_lbl} levelga nechta daily test yozilsin?\n(son kiriting, masalan: 20)",
        reply_markup=cancel_keyboard(lang),
    )


@dp.callback_query(lambda c: c.data == 'teacher_ai_daily_tests_stock_view')
async def handle_teacher_ai_daily_tests_stock_view(callback: CallbackQuery):
    user = get_user_by_telegram(str(callback.from_user.id))
    lang = detect_lang_from_user(user or callback.from_user)
    if not user or user.get("login_type") != 3:
        await callback.answer()
        return
    subject = (user.get("subject") or "").strip().title()
    if subject not in ("English", "Russian"):
        await callback.answer(t(lang, "invalid_subject"), show_alert=True)
        return
    await callback.answer()
    await callback.message.answer(
        _teacher_format_daily_tests_stock_html(lang, int(user["id"]), subject),
        parse_mode="HTML",
    )


@dp.callback_query(lambda c: c.data == 'teacher_ai_daily_tests_history_view')
async def handle_teacher_ai_daily_tests_history_view(callback: CallbackQuery):
    user = get_user_by_telegram(str(callback.from_user.id))
    lang = detect_lang_from_user(user or callback.from_user)
    if not user or user.get("login_type") != 3:
        await callback.answer()
        return
    await callback.answer()
    await callback.message.answer(_teacher_format_daily_tests_history(lang, int(user["id"])), parse_mode="HTML")


@dp.callback_query(lambda c: c.data.startswith("tc_lb_p:"))
async def handle_teacher_dcoin_lb_page(callback: CallbackQuery):
    user = get_user_by_telegram(str(callback.from_user.id))
    lang = detect_lang_from_user(user or callback.from_user)
    if not user or user.get("login_type") != 3:
        await callback.answer()
        return
    subject = (user.get("subject") or "").strip().title()
    if subject not in ("English", "Russian"):
        await callback.answer(t(lang, "invalid_subject"), show_alert=True)
        return
    try:
        page = int(callback.data.split(":", 1)[1])
    except (IndexError, ValueError):
        await callback.answer()
        return
    await callback.answer()
    await _teacher_send_dcoin_leaderboard(callback.message, lang, subject, page)


@dp.callback_query(lambda c: c.data == 'teacher_ai_request_access')
async def handle_teacher_ai_request_access(callback: CallbackQuery):
    user = get_user_by_telegram(str(callback.from_user.id))
    lang = detect_lang_from_user(user or callback.from_user)
    if not user:
        await callback.answer()
        return

    # Notify all admins (best-effort)
    req_teacher_id = user['id']
    req_teacher = user
    req_name = f"{req_teacher.get('first_name','')} {req_teacher.get('last_name','')}".strip()

    kb = InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text=t(lang, 'btn_yes'), callback_data=f'admin_ai_request_approve_{req_teacher_id}')],
            [InlineKeyboardButton(text=t(lang, 'btn_no'), callback_data=f'admin_ai_request_deny_{req_teacher_id}')],
        ]
    )
    for admin_id in ALL_ADMIN_IDS:
        try:
            await bot.send_message(
                int(admin_id),
                f"🤖 AI ruxsati so‘rovi\n\nO'qituvchi: {req_name}\nID: {req_teacher_id}\n\nTasdiqlash uchun tugmani bosing.",
                reply_markup=kb,
            )
        except Exception:
            pass

    await callback.answer()
    await callback.message.answer(t(lang, 'teacher_ai_request_sent'))


@dp.callback_query(lambda c: c.data.startswith("daily_tests_level_"))
async def handle_daily_tests_level_pick(callback: CallbackQuery):
    teacher = get_user_by_telegram(str(callback.from_user.id))
    lang = detect_lang_from_user(teacher or callback.from_user)
    if not teacher or teacher.get('login_type') != 3:
        await callback.answer()
        await callback.message.answer(t(lang, 'please_send_start'))
        return

    if not get_daily_test_upload_permission(teacher['id']):
        await callback.answer(t(lang, 'teacher_daily_tests_permission_denied'), show_alert=True)
        return

    ts = get_teacher_state(callback.message.chat.id)
    subject = (ts.get('data') or {}).get('subject') or (teacher.get('subject') or '').title()
    if subject not in ('English', 'Russian'):
        await callback.answer()
        await callback.message.answer(t(lang, 'invalid_subject'))
        return
    level = callback.data.split("_")[-1].upper()
    allowed_levels = ('A1', 'A2', 'B1', 'B2') if subject == 'Russian' else ('A1', 'A2', 'B1', 'B2', 'C1')
    if level not in allowed_levels:
        await callback.answer()
        return

    ts['data'] = ts.get('data') or {}
    ts['data']['level'] = level
    ts['step'] = 'await_daily_tests_xlsx'

    await callback.message.answer(t(lang, 'teacher_daily_tests_send_xlsx'), reply_markup=cancel_keyboard(lang))
    await callback.answer()



@dp.message(Command('import_vocab'))
async def cmd_import_vocab(message: Message):
    logger.debug(f"cmd_import_vocab: user_id={message.from_user.id}")
    user = get_user_by_telegram(str(message.from_user.id))
    lang = detect_lang_from_user(user or message.from_user)
    if not user or user.get('login_type') != 3:
        await message.answer(t(lang, 'please_send_start'))
        return
    await message.answer(t(lang, 'use_menu_button_vocab'))


@dp.message()
async def handle_teacher_file(message: Message):
    """Handles document upload only when in import flow"""
    state = get_teacher_state(message.chat.id)
    user = get_user_by_telegram(str(message.from_user.id))
    lang = detect_lang_from_user(user or message.from_user)
    if not state.get('step'):
        return

    if state['step'] == 'await_vocab_file':
        subject = state['data'].get('subject')
        # Validate subject
        if subject not in ['English', 'Russian']:
            await message.answer(t(lang, 'invalid_subject'))
            state['step'] = None
            state['data'] = {}
            return

        from vocabulary import import_words_from_excel
        import io

        doc = message.document
        if not doc:
            await message.answer(t(lang, 'only_xlsx_allowed'))
            state['step'] = None
            state['data'] = {}
            return

        file_name = (doc.file_name or "").strip()
        if not file_name.lower().endswith(".xlsx"):
            await message.answer(t(lang, 'only_xlsx_allowed'))
            state['step'] = None
            state['data'] = {}
            return

        bio = io.BytesIO()
        await message.bot.download(file=doc.file_id, destination=bio)
        bio.seek(0)

        # Enforce teacher's subject again for safety
        if user and (user.get('subject') or '').title() != subject.title() and (user.get('subject') or '').title() != 'Both':
            await message.answer(t(lang, 'only_own_subject_allowed'))
            state['step'] = None
            state['data'] = {}
            return

        lang_code = 'en' if subject.lower() == 'english' else 'ru'
        try:
            report = import_words_from_excel(
                bio.read(),
                file_name,
                user['id'],
                subject,
                lang_code,
            )
            await message.answer(
                t(
                    lang,
                    'teacher_import_done_report',
                    inserted=report['inserted'],
                    skipped=report['skipped'],
                    duplicates=', '.join(report.get('duplicates', [])[:5]) if report.get('duplicates') else t(lang, 'none_short'),
                )
            )
        except Exception as e:
            await message.answer(t(lang, 'error_with_reason', error=str(e)))

        state['step'] = None
        state['data'] = {}
        return

    if state['step'] == 'await_daily_tests_xlsx':
        subject = (state['data'] or {}).get('subject')
        level = (state['data'] or {}).get('level')

        allowed_levels = ('A1', 'A2', 'B1', 'B2') if subject == 'Russian' else ('A1', 'A2', 'B1', 'B2', 'C1')
        if subject not in ('English', 'Russian') or level not in allowed_levels:
            await message.answer(t(lang, 'teacher_daily_tests_upload_error', error='Invalid subject/level state'))
            state['step'] = None
            state['data'] = {}
            return

        doc = message.document
        if not doc:
            await message.answer(t(lang, 'only_xlsx_allowed'))
            state['step'] = None
            state['data'] = {}
            return

        file_name = (doc.file_name or '').strip()
        if not file_name.lower().endswith('.xlsx'):
            await message.answer(t(lang, 'only_xlsx_allowed'))
            state['step'] = None
            state['data'] = {}
            return

        # Enforce permission again for safety
        if not get_daily_test_upload_permission(user['id']):
            await message.answer(t(lang, 'teacher_daily_tests_permission_denied'))
            state['step'] = None
            state['data'] = {}
            return

        from daily_tests import import_daily_tests_from_xlsx
        import io

        bio = io.BytesIO()
        await message.bot.download(file=doc.file_id, destination=bio)
        bio.seek(0)

        try:
            report = import_daily_tests_from_xlsx(
                bio.read(),
                file_name,
                user['id'],
                subject,
                level,
            )
            await message.answer(
                t(
                    lang,
                    'teacher_daily_tests_upload_done',
                    inserted=(report.get('inserted', 0) if isinstance(report, dict) else getattr(report, "inserted", 0)),
                    skipped=(report.get('skipped', 0) if isinstance(report, dict) else getattr(report, "skipped", 0)),
                )
            )
        except Exception as e:
            await message.answer(t(lang, 'teacher_daily_tests_upload_error', error=str(e)))

        state['step'] = None
        state['data'] = {}
        return

    if state['step'] == 'teacher_arena_excel_file':
        gid = int((state.get('data') or {}).get('arena_group_id') or 0)
        group = get_group(gid)
        if not group:
            await message.answer(t(lang, "teacher_group_arena_group_not_found"))
            state['step'] = None
            state['data'] = {}
            return

        # Determine subject/level from group.
        arena_subject = (group.get("subject") or user.get("subject") or "English")
        arena_subject = str(arena_subject).strip().title()
        if arena_subject not in ("English", "Russian"):
            arena_subject = "English"

        arena_level = (group.get("level") or "A1").upper()
        if arena_level not in ("A1", "A2", "B1", "B2", "C1"):
            arena_level = "A1"

        doc = message.document
        if not doc:
            await message.answer(t(lang, 'only_xlsx_allowed'))
            state['step'] = None
            state['data'] = {}
            return

        file_name = (doc.file_name or '').strip()
        if not file_name.lower().endswith('.xlsx'):
            await message.answer(t(lang, 'only_xlsx_allowed'))
            state['step'] = None
            state['data'] = {}
            return

        # Enforce permission again for safety (same as daily tests import).
        if not get_daily_test_upload_permission(user['id']):
            await message.answer(t(lang, 'teacher_daily_tests_permission_denied'))
            state['step'] = None
            state['data'] = {}
            return

        from daily_tests import import_daily_tests_from_xlsx
        import io

        bio = io.BytesIO()
        await message.bot.download(file=doc.file_id, destination=bio)
        bio.seek(0)

        start_ts = datetime.utcnow()
        try:
            report = import_daily_tests_from_xlsx(
                bio.read(),
                file_name,
                user['id'],
                arena_subject,
                arena_level,
            )
        except Exception as e:
            await message.answer(t(lang, 'teacher_daily_tests_upload_error', error=str(e)))
            state['step'] = None
            state['data'] = {}
            return

        inserted = int((report.get('inserted', 0) if isinstance(report, dict) else getattr(report, "inserted", 0)) or 0)

        # Select newly imported rows (by created_at window).
        conn = get_conn()
        cur = conn.cursor()
        try:
            cur.execute(
                '''
                SELECT id
                FROM daily_tests_bank
                WHERE created_by=? AND subject=? AND level=?
                  AND active=1 AND created_at >= ?
                ORDER BY created_at ASC, id ASC
                LIMIT ?
                ''',
                (user['id'], arena_subject, arena_level, start_ts, max(1, inserted)),
            )
            rows = cur.fetchall()
            bank_ids = [int(r['id']) for r in rows]
        finally:
            conn.close()

        arena_bank_ids = copy_daily_tests_bank_rows_to_arena_questions(
            bank_ids=bank_ids,
            created_by=user['id'],
        )

        if not arena_bank_ids:
            await message.answer(t(lang, "teacher_group_arena_generation_failed"))
            state['step'] = None
            state['data'] = {}
            return

        session_id = create_arena_group_session(
            group_id=gid,
            subject=arena_subject,
            level=arena_level,
            question_count=len(arena_bank_ids),
            bank_ids=arena_bank_ids,
            created_by_teacher_id=user['id'],
        )

        state['step'] = None
        state['data'] = {}

        await message.answer(
            t(
                lang,
                "teacher_group_arena_ready",
                session_id=session_id,
                question_count=len(arena_bank_ids),
            ),
            reply_markup=InlineKeyboardMarkup(
                inline_keyboard=[
                    [InlineKeyboardButton(text=t(lang, "teacher_group_arena_send_questions_btn"), callback_data=f"teacher_arena_send:{session_id}")]
                ]
            ),
        )
        return


@dp.message(Command('export_vocab'))
async def cmd_export_vocab_teacher(message: Message):
    user = get_user_by_telegram(str(message.from_user.id))
    lang = detect_lang_from_user(user or message.from_user)
    if not user or user.get('login_type') != 3:
        await message.answer(t(lang, 'please_send_start'))
        return
    await message.answer(t(lang, 'use_menu_button_vocab'))


@dp.callback_query(lambda c: c.data.startswith('attendance_mark_'))
async def handle_callbacks(callback: CallbackQuery):
    data = callback.data
    
    if data.startswith('attendance_mark_'):
        parts = data.split('_')
        user_id = int(parts[2])
        group_id = int(parts[3])
        status = parts[4]  # Present/Absent
        
        user = get_user_by_telegram(str(callback.from_user.id))
        if not user or user.get("login_type") != 3:
            await callback.answer()
            return
        group = get_group(group_id)
        if not _teacher_can_manage_group(user, group):
            await callback.answer(t(detect_lang_from_user(user or callback.from_user), 'no_permission_for_group'), show_alert=True)
            return
        
        today = datetime.utcnow().strftime('%Y-%m-%d')
        add_attendance(user_id, group_id, today, status)
        
        lang = detect_lang_from_user(get_user_by_telegram(str(user_id)) or callback.from_user)
        await callback.message.answer(t(lang, 'davomat_marked', status=status))
        await callback.answer()
        return


# Subject selection callback is no longer needed since "Both" is removed


@dp.callback_query(lambda c: c.data == "teacher_arena_rules")
async def handle_teacher_arena_rules(callback: CallbackQuery):
    await callback.answer()
    lang = detect_lang_from_user(callback.from_user)
    await callback.message.answer(t(lang, "teacher_arena_rules_text"))


@dp.callback_query(lambda c: c.data == "teacher_arena_group_start")
async def handle_teacher_arena_group_start(callback: CallbackQuery):
    user = get_user_by_telegram(str(callback.from_user.id))
    lang = detect_lang_from_user(user or callback.from_user)
    if not user or user.get("login_type") != 3:
        await callback.answer()
        await callback.message.answer(t(lang, 'please_send_start'))
        return
    groups = _manageable_groups_for_teacher(user['id']) or []
    if not groups:
        await callback.answer()
        await callback.message.answer(t(lang, 'groups_not_found'))
        return
    groups.sort(key=lambda g: (g.get("name") or ""))
    total_pages = max(1, (len(groups) + 9) // 10)
    page = 0
    chunk = _teacher_arena_groups_page(groups, page, 10)
    text = _render_teacher_arena_group_list_text(lang, chunk, page, total_pages)
    kb = _teacher_arena_group_pick_keyboard(chunk, page, total_pages, lang)
    await callback.answer()
    await callback.message.answer(text, reply_markup=kb, parse_mode="HTML")


@dp.callback_query(lambda c: (c.data or "").startswith("teacher_arena_grp:"))
async def handle_teacher_arena_group_list_nav(callback: CallbackQuery):
    user = get_user_by_telegram(str(callback.from_user.id))
    lang = detect_lang_from_user(user or callback.from_user)
    if not user or user.get("login_type") != 3:
        await callback.answer()
        return
    parts = (callback.data or "").split(":")
    if len(parts) != 3:
        await callback.answer(t(lang, "invalid_action"), show_alert=True)
        return
    action, val = parts[1], parts[2]
    groups = _manageable_groups_for_teacher(user['id']) or []
    groups.sort(key=lambda g: (g.get("name") or ""))
    total_pages = max(1, (len(groups) + 9) // 10)
    if action == "page":
        try:
            page = int(val)
        except ValueError:
            await callback.answer(t(lang, "invalid_action"), show_alert=True)
            return
        page = max(0, min(page, total_pages - 1))
        chunk = _teacher_arena_groups_page(groups, page, 10)
        text = _render_teacher_arena_group_list_text(lang, chunk, page, total_pages)
        kb = _teacher_arena_group_pick_keyboard(chunk, page, total_pages, lang)
        try:
            await callback.message.edit_text(text, reply_markup=kb, parse_mode="HTML")
        except Exception:
            await callback.message.answer(text, reply_markup=kb, parse_mode="HTML")
        await callback.answer()
        return
    if action == "pick":
        try:
            gid = int(val)
        except ValueError:
            await callback.answer(t(lang, "invalid_action"), show_alert=True)
            return
        group = get_group(gid)
        if not _teacher_can_manage_group(user or {}, group):
            await callback.answer(t(lang, "teacher_group_arena_permission_denied"), show_alert=True)
            return
        kb = InlineKeyboardMarkup(
            inline_keyboard=[
                [InlineKeyboardButton(text=t(lang, "teacher_group_arena_source_ai_btn"), callback_data=f"teacher_arena_source_ai_{gid}")],
                [InlineKeyboardButton(text=t(lang, "teacher_group_arena_source_excel_btn"), callback_data=f"teacher_arena_source_excel_{gid}")],
            ]
        )
        await callback.answer()
        await callback.message.answer(t(lang, "teacher_group_arena_source_pick"), reply_markup=kb)
        return
    await callback.answer(t(lang, "invalid_action"), show_alert=True)


@dp.callback_query(lambda c: c.data.startswith("teacher_arena_source_excel_"))
async def handle_teacher_arena_source_excel(callback: CallbackQuery):
    gid = int(callback.data.split("_")[-1])
    ts = get_teacher_state(callback.message.chat.id)
    ts["step"] = "teacher_arena_excel_file"
    ts["data"] = {"arena_group_id": gid}
    lang = detect_lang_from_user(callback.from_user)
    await callback.answer()
    await callback.message.answer(
        t(lang, "teacher_group_arena_excel_mode_selected", group_id=gid),
        reply_markup=cancel_keyboard(lang),
    )


@dp.callback_query(lambda c: c.data.startswith("teacher_arena_source_ai_"))
async def handle_teacher_arena_source_ai(callback: CallbackQuery):
    gid = int(callback.data.split("_")[-1])
    ts = get_teacher_state(callback.message.chat.id)
    ts['step'] = 'teacher_arena_ai_topic_mode'
    ts['data'] = {'arena_group_id': gid}

    kb = InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text=t(detect_lang_from_user(callback.from_user), "teacher_group_arena_ai_topic_current_btn"), callback_data=f"teacher_arena_ai_topic:{gid}:current")],
            [InlineKeyboardButton(text=t(detect_lang_from_user(callback.from_user), "teacher_group_arena_ai_topic_any_btn"), callback_data=f"teacher_arena_ai_topic:{gid}:any")],
        ]
    )
    await callback.answer()
    await callback.message.answer(t(detect_lang_from_user(callback.from_user), "teacher_group_arena_ai_topic_mode_pick"), reply_markup=kb)


@dp.callback_query(lambda c: c.data.startswith("teacher_arena_ai_topic:"))
async def handle_teacher_arena_ai_topic_mode(callback: CallbackQuery):
    """callback_data: teacher_arena_ai_topic:{gid}:current|any (3 ta qism)."""
    parts = (callback.data or "").split(":")
    if len(parts) != 3:
        await callback.answer(t(detect_lang_from_user(callback.from_user), "invalid_action"), show_alert=True)
        return
    _, gid_str, mode = parts
    mode = (mode or "").strip().lower()
    if mode not in {"current", "any"}:
        await callback.answer(t(detect_lang_from_user(callback.from_user), "invalid_action"), show_alert=True)
        return
    try:
        gid = int(gid_str)
    except ValueError:
        await callback.answer(t(detect_lang_from_user(callback.from_user), "invalid_action"), show_alert=True)
        return

    ts = get_teacher_state(callback.message.chat.id)
    ts["data"] = ts.get("data") or {}
    ts["data"]["arena_group_id"] = gid
    ts["data"]["topic_mode"] = mode
    ts["step"] = "teacher_arena_ai_count"

    lang = detect_lang_from_user(callback.from_user)
    await callback.answer()
    await callback.message.answer(
        t(lang, "teacher_group_arena_ai_count_prompt", example=10),
        reply_markup=cancel_keyboard(lang),
    )


@dp.callback_query(lambda c: c.data.startswith("teacher_arena_send:"))
async def handle_teacher_arena_send_questions(callback: CallbackQuery):
    """
    Teacher-generated Group Arena savollarini present bo'lgan studentlarga yuboradi.
    Studentlar `arena_join_direct:group:{session_id}` bosganda quiz boshlanadi.
    """
    session_id = int(callback.data.split(":", 1)[-1])

    teacher = get_user_by_telegram(str(callback.from_user.id))
    lang = detect_lang_from_user(teacher or callback.from_user)

    session = get_arena_group_session(session_id)
    if not session:
        await callback.answer(t(lang, "teacher_group_arena_session_not_found"), show_alert=True)
        return
    if (session.get("created_by_teacher_id") or None) != (teacher.get("id") if teacher else None):
        # Safety: only creating teacher can send
        await callback.answer(t(lang, 'no_permission_for_group'), show_alert=True)
        return

    gid = int(session["group_id"])
    group = get_group(gid)
    if not _teacher_can_manage_group(teacher or {}, group):
        await callback.answer(t(lang, 'no_permission_for_group'), show_alert=True)
        return

    if not is_group_lesson_window_active(gid):
        await callback.answer(t(lang, "arena_group_not_lesson_time"), show_alert=True)
        return

    tz = pytz.timezone("Asia/Tashkent")
    today = datetime.now(tz).strftime("%Y-%m-%d")
    present_students = get_present_students_for_group_date(gid, today) or []

    joined = 0
    notify_bot = student_notify_bot or bot
    for st in present_students:
        tg = st.get("telegram_id")
        if not tg:
            continue
        try:
            await notify_bot.send_message(
                int(tg),
                t(
                    detect_lang_from_user(st),
                    "teacher_group_arena_student_notification",
                    count=session.get('question_count'),
                    sec=40,
                ),
                reply_markup=InlineKeyboardMarkup(
                    inline_keyboard=[
                        [InlineKeyboardButton(text=t(detect_lang_from_user(st), "arena_join_free_btn"), callback_data=f"arena_join_direct:group:{session_id}")]
                    ]
                ),
            )
            joined += 1
        except Exception:
            pass

    # Expected players: number of students who are marked present for this group/date.
    set_arena_group_session_expected_players(session_id, expected_players=len(present_students))

    # Snapshot the selected questions into tmp table for delayed promotion.
    try:
        from db import populate_arena_group_questions_tmp

        populate_arena_group_questions_tmp(session_id)
    except Exception:
        logger.exception("populate_arena_group_questions_tmp failed session_id=%s", session_id)

    await callback.answer()
    await callback.message.answer(t(lang, "teacher_group_arena_send_done", joined=joined))

    try:
        xlsx_bytes = _build_group_arena_questions_xlsx(session_id)
        await callback.message.answer_document(
            BufferedInputFile(xlsx_bytes, filename=f"group_arena_{session_id}_questions.xlsx"),
            caption=t(lang, "teacher_group_arena_questions_xlsx_caption"),
        )
    except Exception:
        logger.exception("teacher arena questions xlsx failed session_id=%s", session_id)

    try:
        snap = get_group_arena_teacher_snapshot(session_id)
        if snap:
            live_text = _format_teacher_arena_live_text(snap, lang)
            sent = await callback.message.answer(live_text, parse_mode="HTML")
            set_arena_group_session_teacher_message(session_id, callback.message.chat.id, sent.message_id)
            enqueue_arena_group_teacher_refresh(session_id)
    except Exception:
        logger.exception("teacher arena live message failed session_id=%s", session_id)


@dp.callback_query(lambda c: (c.data or "").startswith("teacher_arena_matrix_dl:"))
async def handle_teacher_arena_matrix_download(callback: CallbackQuery):
    teacher = get_user_by_telegram(str(callback.from_user.id))
    lang = detect_lang_from_user(teacher or callback.from_user)
    try:
        session_id = int((callback.data or "").split(":", 1)[-1])
    except ValueError:
        await callback.answer(t(lang, "invalid_action"), show_alert=True)
        return
    session = get_arena_group_session(session_id)
    if not session:
        await callback.answer(t(lang, "teacher_group_arena_session_not_found"), show_alert=True)
        return
    if (session.get("created_by_teacher_id") or None) != (teacher.get("id") if teacher else None):
        await callback.answer(t(lang, "no_permission_for_group"), show_alert=True)
        return
    try:
        raw = _build_group_arena_matrix_xlsx(session_id)
        await callback.message.answer_document(
            BufferedInputFile(raw, filename=f"group_arena_{session_id}_matrix.xlsx"),
            caption=t(lang, "teacher_group_arena_matrix_xlsx_caption"),
        )
    except Exception:
        logger.exception("matrix xlsx failed session_id=%s", session_id)
        await callback.message.answer(t(lang, "system_error_try_later_contact_admin"))
    await callback.answer()


async def run_teacher_bot():
    global bot, student_notify_bot
    print("[STARTUP] teacher_bot run_teacher_bot() starting")
    if not TEACHER_BOT_TOKEN:
        raise RuntimeError("TEACHER_BOT_TOKEN is not set. Put it in .env (TEACHER_BOT_TOKEN=...) and retry.")
    bot = create_resilient_bot(TEACHER_BOT_TOKEN)

    try:
        ensure_subject_dcoin_schema()
        ensure_dcoin_schema_migrations()
    except Exception:
        logger.exception("Teacher bot startup: D'coin schema ensure failed")
    if not validate_dcoin_runtime_ready(context="teacher_bot startup"):
        raise RuntimeError("Teacher bot startup aborted: D'coin runtime schema is not ready")
    print("[STARTUP] teacher_bot validate_dcoin_runtime_ready() ok")

    if not STUDENT_BOT_TOKEN:
        raise RuntimeError("STUDENT_BOT_TOKEN is not set. Put it in .env (STUDENT_BOT_TOKEN=...) and retry.")
    student_notify_bot = create_resilient_bot(STUDENT_BOT_TOKEN)
    set_attendance_student_notify_bot(student_notify_bot)

    # Restore sessions on startup
    try:
        restored_count = restore_sessions_on_startup()
        logger.info(f"🔄 Restored {restored_count} teacher sessions on startup")
    except Exception as e:
        logger.error(f"Failed to restore sessions: {e}")
    
    from config import ADMIN_CHAT_IDS
    spawn_guarded_task(attendance_scheduler(bot, role="teacher", admin_chat_ids=ADMIN_CHAT_IDS), name="teacher_attendance_scheduler")
    spawn_guarded_task(teacher_arena_refresh_worker(bot), name="teacher_arena_refresh_worker")

    logger.info("Teacher bot started successfully")
    await run_bot_dispatcher(dp=dp, bot=bot, bot_name="teacher", webhook_port=TEACHER_WEBHOOK_PORT)


@dp.error()
async def error_handler(event: types.ErrorEvent):
    exc = event.exception
    logger.error(f"Error in teacher bot: {exc}", exc_info=True)

    # 1) Write to log.txt
    try:
        log_file = Path(__file__).resolve().parent / "log.txt"
        stack = "".join(traceback.format_exception(type(exc), exc, exc.__traceback__))
        with open(log_file, "a", encoding="utf-8") as f:
            f.write("\n" + "=" * 80 + "\n")
            f.write(f"Teacher bot error: {repr(exc)}\n")
            f.write(stack)
            f.write("\n")
    except Exception:
        logger.exception("Failed to write teacher bot error to log.txt")

    # 2) Notify admins (best-effort)
    try:
        if bot:
            update_info = getattr(event, "update", None)
            update_id = getattr(update_info, "update_id", None)
            short = f"Teacher bot xatolik: {repr(exc)}"
            if update_id is not None:
                short += f" (update_id={update_id})"
            for aid in ADMIN_CHAT_IDS:
                await bot.send_message(int(aid), short)
    except Exception:
        logger.exception("Failed to notify admins about teacher bot error")


@dp.callback_query(lambda c: c.data.startswith('set_lang_me_'))
async def handle_set_lang_me_teacher(callback: CallbackQuery):
    code = callback.data.split('_')[-1]
    from db import set_user_language_by_telegram
    ok = set_user_language_by_telegram(str(callback.from_user.id), code)
    if ok:
        await callback.answer()
        await callback.message.answer(t(code, 'lang_set'))
        # Send updated main menu in new language
        await callback.message.answer(t(code, 'select_from_menu'), reply_markup=teacher_main_keyboard(code))
    else:
        await callback.answer()
        await callback.message.answer(t(code, 'please_send_start'))
