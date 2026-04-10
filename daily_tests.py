import io
from typing import Any, Iterable

from openpyxl import load_workbook


def _normalize_str(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _infer_header_row(raw_row: list[Any]) -> bool:
    """
    Heuristic: if the first non-empty cell looks like a header, skip the row.
    """
    first = _normalize_str(raw_row[0]).lower()
    if not first:
        return False
    return first in {"question", "savol", "q", "questions"} or "question" in first


def _infer_correct_option_index(
    correct_raw: Any,
    options: list[str],
) -> int:
    """
    Resolve correct option index (1..4).
    Supports:
    - F column as integer 1..4
    - F column as exact option text match (case-insensitive)
    - option markers: '✅', '(correct)', or leading '*'
    """
    # 1) Explicit F column
    if correct_raw is not None and str(correct_raw).strip() != "":
        cr = str(correct_raw).strip()
        try:
            idx = int(cr)
            if 1 <= idx <= 4:
                return idx
        except Exception:
            pass

        cr_low = cr.lower()
        for i, opt in enumerate(options, start=1):
            if opt.lower().strip() == cr_low:
                return i

    # 2) Markers inside options
    markers = []
    for i, opt in enumerate(options, start=1):
        low = opt.lower()
        is_marked = any(
            [
                low.startswith("✅"),
                "(correct)" in low,
                low.startswith("*"),
                "correct" in low,
            ]
        )
        if is_marked:
            markers.append(i)

    if len(markers) == 1:
        return markers[0]

    raise ValueError("Could not determine correct answer. Provide it in column F (1-4 or exact text).")


def parse_daily_tests_xlsx(file_bytes: bytes) -> list[dict[str, Any]]:
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True)
    sheet = wb.active

    rows_iter = sheet.iter_rows(values_only=True)
    items: list[dict[str, Any]] = []

    for row_idx, row in enumerate(rows_iter, start=1):
        row = list(row or [])
        # Columns A..F are used: A=question, B-E=4 options, F=correct
        while len(row) < 6:
            row.append(None)

        # Skip empty rows
        if all(v is None or str(v).strip() == "" for v in row[:5]):
            continue

        if row_idx == 1 and _infer_header_row(row):
            continue

        question = _normalize_str(row[0])
        opt1 = _normalize_str(row[1])
        opt2 = _normalize_str(row[2])
        opt3 = _normalize_str(row[3])
        opt4 = _normalize_str(row[4])
        correct_raw = row[5]

        options = [opt1, opt2, opt3, opt4]
        if not question or any(not o for o in options):
            # Incomplete row
            continue

        correct_idx = _infer_correct_option_index(correct_raw, options)

        items.append(
            {
                "question": question,
                "options": options,
                "correct_option_index": correct_idx,
            }
        )

    return items


def import_daily_tests_from_xlsx(
    file_bytes: bytes,
    file_name: str,
    teacher_id: int,
    subject: str,
    level: str,
    *,
    batch_size: int = 500,
) -> dict[str, Any]:
    """
    Import daily tests bank items.

    Excel format:
    - Column A: question
    - Column B-E: 4 answer options
    - Column F: correct answer (either index 1..4 or exact text from B-E; or marked option text).
    """
    items = parse_daily_tests_xlsx(file_bytes)
    if not items:
        raise ValueError("No valid daily test rows found in XLSX.")

    inserted = 0
    skipped = 0
    errors: list[str] = []

    from db import get_conn, DB_WRITE_LOCK

    conn = get_conn()
    cur = conn.cursor()

    # Basic de-dup within the uploaded file
    seen_keys: set[tuple[str, tuple[str, ...], int]] = set()

    def iter_batches() -> Iterable[list[tuple[Any, ...]]]:
        nonlocal skipped, inserted
        batch: list[tuple[Any, ...]] = []
        for it in items:
            q = it["question"]
            opts = tuple(it["options"])
            correct_idx = int(it["correct_option_index"])
            key = (q.lower().strip(), opts, correct_idx)
            if key in seen_keys:
                skipped += 1
                continue
            seen_keys.add(key)
            batch.append(
                (
                    teacher_id,
                    subject,
                    level,
                    q,
                    opts[0],
                    opts[1],
                    opts[2],
                    opts[3],
                    correct_idx,
                    1,  # active
                )
            )
            if len(batch) >= batch_size:
                yield batch
                batch = []
        if batch:
            yield batch

    insert_sql = '''
        INSERT INTO daily_tests_bank
        (created_by, subject, level, question, option_a, option_b, option_c, option_d, correct_option_index, active)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    '''

    try:
        with DB_WRITE_LOCK:
            for batch in iter_batches():
                cur.executemany(insert_sql, batch)
                inserted += len(batch)
            conn.commit()
    except Exception as e:
        conn.rollback()
        errors.append(str(e))
        raise
    finally:
        conn.close()

    # errors are not expected because we re-raise
    return {"file_name": file_name, "inserted": inserted, "skipped": skipped, "errors": errors[:3]}

